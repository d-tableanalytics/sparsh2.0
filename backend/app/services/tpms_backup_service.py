"""
TPMS export / import — the whole module as one .xlsx workbook, a sheet per collection.

Two jobs in one file because they are two directions of one format: whatever `export_workbook`
writes, `import_workbook` must be able to read back. Keeping the column encoding (`_cell`) and
its inverse (`_value`) side by side is what stops the pair drifting apart.

The workbook is a working document, not only a backup. The `Schedules` sheet leads, carries
human-readable columns (activity names, company names, member EMAILS — never raw ObjectIds in
the fields you are expected to type), and is meant to be extended: add rows underneath the
exported ones, re-import, and each new row becomes a real scheduled activity.

Two import paths, deliberately different:

  • `Schedules`  — a new row is NOT inserted as a document. It is replayed through
    `tpms_schedule_service.create_schedule`, the same function the Schedule modal calls. So a
    bulk-loaded activity expands its recurrence, attaches catalogue reminders, writes its
    Activity_Tracker rows, sends the schedule mail and mints the per-assignee form links —
    identical to one created by hand. There is no second scheduling code path to keep in sync.

  • every other sheet — add-only restore. A row whose `_id` already exists is skipped; nothing
    is ever updated or deleted. Re-running an import is therefore harmless, and a stale file
    cannot roll back live data.

Admin-only at the route: the workbook contains live form-link tokens and respondent email
addresses, so it must never reach a client-side user.
"""
from __future__ import annotations

import io
import json
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Tuple

from bson import ObjectId
from bson.errors import InvalidId
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.db.mongodb import get_collection
from app.models.tpms import (
    COLL_ACTION_ITEMS,
    COLL_ACTIVITIES,
    COLL_ACTIVITY_TRACKER,
    COLL_DEPARTMENTS,
    COLL_ESCALATIONS,
    COLL_MAIL_TEMPLATES,
    COLL_MIGRATION_MAP,
    COLL_REMINDER_LOGS,
    COLL_REMINDER_RULES,
    COLL_RESCHEDULE_REQUESTS,
    COLL_SUCCESS_MEASURES,
    COLL_TASK_UPLOADS,
    COLL_WHATSAPP_TEMPLATES,
    TPMS_EVENT_KIND,
)
from app.utils.calendar_utils import CALENDAR_COLLECTIONS

logger = logging.getLogger(__name__)

# Schedules live in the ERP calendar collections tagged `kind: "tpms_activity"` (see
# calendar_utils), not in a table of their own — so they are gathered by scan, not by name.
SCHEDULE_COLLECTIONS = CALENDAR_COLLECTIONS + ["calendar_events"]

# Every TPMS-owned collection, in the order the sheets appear. `tpms_form_assignments` is a
# string literal rather than a constant because it belongs to tpms_form_link_service, which
# defines it locally; importing that module here would be a cycle at call time.
BACKUP_COLLECTIONS: List[str] = [
    COLL_ACTIVITIES,
    COLL_DEPARTMENTS,
    COLL_REMINDER_RULES,
    COLL_MAIL_TEMPLATES,
    COLL_WHATSAPP_TEMPLATES,
    COLL_RESCHEDULE_REQUESTS,
    COLL_ACTIVITY_TRACKER,
    COLL_ESCALATIONS,
    COLL_ACTION_ITEMS,
    COLL_SUCCESS_MEASURES,
    COLL_REMINDER_LOGS,
    COLL_TASK_UPLOADS,
    COLL_MIGRATION_MAP,
    "tpms_form_assignments",
]

SCHEDULE_SHEET = "Schedules"

# The fillable sheet's columns. The first three are produced by Export and read by Import only
# to decide "existing or new" — a row you ADD leaves Schedule ID blank, and that blank is the
# entire signal that it should be created.
SCHEDULE_COLUMNS: List[str] = [
    "Schedule ID",          # blank ⇒ create this row; filled ⇒ already exists, skipped
    "Batch ID",             # export-only, groups one recurrence expansion
    "Status",               # export-only (Scheduled / Completed / …)
    "Title",
    "Activity",
    "Company Name",
    "Company ID",
    "Plan Date",            # YYYY-MM-DD
    "Time",                 # HH:MM
    "Recurrence",           # One-time | Weekly | Monthly | Periodically
    "Plan End",             # YYYY-MM-DD, required when Recurrence ≠ One-time
    "Weekdays",             # "Mon,Wed" or "0,2" — only for Periodically
    "Departments",          # comma-separated, e.g. "HOD,MD"
    "Company Assigners",    # comma-separated EMAILS of the doers
    "Staff Assigners",      # comma-separated EMAILS of internal SMOps
    "Comment",
]

# A one-line reminder of the contract, written into the sheet so the person filling it in does
# not have to be told separately.
SCHEDULE_HINT = (
    "Add new activities as rows below. Leave 'Schedule ID' EMPTY for a new row — rows that "
    "already have an ID are skipped on import. Company/Staff Assigners are email addresses, "
    "comma-separated. Dates are YYYY-MM-DD."
)

# Excel refuses to store a cell longer than this, so oversized JSON is truncated with a marker
# rather than blowing up the whole export.
_MAX_CELL = 32000

_WEEKDAY_NAMES = {
    "mon": 0, "monday": 0, "tue": 1, "tues": 1, "tuesday": 1, "wed": 2, "weds": 2,
    "wednesday": 2, "thu": 3, "thur": 3, "thurs": 3, "thursday": 3, "fri": 4, "friday": 4,
    "sat": 5, "saturday": 5, "sun": 6, "sunday": 6,
}


# ─────────────────────────────────────────────────────────────
# Cell encoding — and its exact inverse
# ─────────────────────────────────────────────────────────────
def _cell(value: Any) -> Any:
    """A Mongo value as something a spreadsheet cell can hold.

    Lists and dicts become JSON so they survive the round trip intact; `_value` parses them
    back. Everything else is flattened to a string, because a workbook has no richer types.
    """
    if value is None:
        return ""
    if isinstance(value, ObjectId):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        # Written as text, not a native bool: openpyxl round-trips TRUE/FALSE inconsistently
        # across Excel locales, and `_value` maps these two spellings back explicitly.
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float, str)):
        text = value if not isinstance(value, str) else value
        if isinstance(text, str) and len(text) > _MAX_CELL:
            return text[:_MAX_CELL] + "…[truncated]"
        return text
    try:
        dumped = json.dumps(value, default=str)
    except Exception:
        dumped = str(value)
    return dumped[:_MAX_CELL] if len(dumped) > _MAX_CELL else dumped


def _value(raw: Any) -> Any:
    """A cell back into a Mongo value. The inverse of `_cell`.

    Blank means "absent" rather than empty-string: an untouched cell in a spreadsheet carries no
    intent, and writing "" into every unfilled field would overwrite meaning with noise.
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float, bool, datetime, date)):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    if text in ("TRUE", "FALSE"):
        return text == "TRUE"
    if text[0] in "[{" and text[-1] in "]}":
        try:
            return json.loads(text)
        except Exception:
            return text
    # An ISO timestamp written by `_cell` — restored so date queries keep working after a
    # restore. Plain dates like "2026-08-01" are left as text: they are usually a period or a
    # plan date the rest of TPMS reads as a string.
    if len(text) >= 19 and text[4] == "-" and text[7] == "-" and text[10] in "T ":
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            return text
    return text


def _sheet_title(name: str) -> str:
    """Excel caps a sheet name at 31 characters and forbids []:*?/\\ ."""
    clean = "".join("_" if ch in "[]:*?/\\" else ch for ch in name)
    return clean[:31]


def _text(row: Dict[str, Any], key: str) -> str:
    """A schedule column as trimmed text, whatever type the cell came back as."""
    raw = row.get(key)
    if raw is None:
        return ""
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return str(raw).strip()


def _csv(row: Dict[str, Any], key: str) -> List[str]:
    """A comma/semicolon/newline separated column as a clean list."""
    text = _text(row, key)
    if not text:
        return []
    parts: List[str] = []
    for chunk in text.replace(";", ",").replace("\n", ",").split(","):
        item = chunk.strip()
        if item:
            parts.append(item)
    return parts


def _date_text(row: Dict[str, Any], key: str) -> str:
    """A date column as YYYY-MM-DD. Excel hands back a datetime for a real date cell and a
    string for a typed one, so both spellings have to land on the same output."""
    raw = row.get(key)
    if isinstance(raw, (datetime, date)):
        return (raw.date() if isinstance(raw, datetime) else raw).isoformat()
    text = _text(row, key)
    if not text:
        return ""
    # Tolerate the DD-MM-YYYY / DD/MM/YYYY people naturally type.
    for sep in ("-", "/"):
        if text.count(sep) == 2:
            a, b, c = text.split(sep)
            if len(a) == 4:
                return f"{a}-{b.zfill(2)}-{c.zfill(2)}"
            if len(c) == 4:
                return f"{c}-{b.zfill(2)}-{a.zfill(2)}"
    return text


def _time_text(row: Dict[str, Any], key: str) -> str:
    """A time column as HH:MM. Excel may return a datetime/time for a formatted cell."""
    raw = row.get(key)
    if isinstance(raw, datetime):
        return raw.strftime("%H:%M")
    if hasattr(raw, "hour") and hasattr(raw, "minute"):
        return f"{raw.hour:02d}:{raw.minute:02d}"
    text = _text(row, key)
    return text[:5] if text else ""


def _weekdays(row: Dict[str, Any]) -> List[int]:
    """"Mon,Wed" or "0,2" → [0, 2]. Both spellings are accepted because the export writes
    names (readable) while the API and the Schedule modal speak indices."""
    out: List[int] = []
    for token in _csv(row, "Weekdays"):
        key = token.strip().lower()
        if key in _WEEKDAY_NAMES:
            out.append(_WEEKDAY_NAMES[key])
            continue
        try:
            idx = int(float(key))
        except ValueError:
            continue
        if 0 <= idx <= 6:
            out.append(idx)
    return sorted(set(out))


# ─────────────────────────────────────────────────────────────
# Export
# ─────────────────────────────────────────────────────────────
async def _directory() -> Tuple[Dict[str, str], Dict[str, str]]:
    """(user id → email, company id → name) for the whole instance.

    Loaded once and reused for every schedule row. The alternative — resolving names per event —
    is a query per assignee per row, which on a full export is thousands of round trips.
    """
    people: Dict[str, str] = {}
    for coll in ("learners", "staff"):
        for user in await get_collection(coll).find(
            {}, {"email": 1, "full_name": 1, "first_name": 1, "last_name": 1}
        ).to_list(20000):
            label = user.get("email") or user.get("full_name") or ""
            if label:
                people[str(user["_id"])] = label

    companies: Dict[str, str] = {}
    for company in await get_collection("companies").find({}, {"name": 1}).to_list(5000):
        companies[str(company["_id"])] = company.get("name") or ""
    return people, companies


async def _schedule_rows() -> List[List[Any]]:
    """Every TPMS activity across the calendar collections, as fillable sheet rows."""
    people, companies = await _directory()
    rows: List[List[Any]] = []

    for coll_name in SCHEDULE_COLLECTIONS:
        try:
            events = await get_collection(coll_name).find(
                {"kind": TPMS_EVENT_KIND}).to_list(20000)
        except Exception as exc:
            logger.warning("TPMS export: could not read %s — %s", coll_name, exc)
            continue

        for event in events:
            meta = event.get("activity_meta") or {}
            start = str(event.get("start") or "")
            company_id = str(event.get("company_id") or "")
            rows.append([
                str(event.get("_id")),
                str(event.get("tpms_batch_id") or ""),
                event.get("tpms_status") or "",
                event.get("title") or "",
                event.get("activity") or "",
                event.get("company_name") or companies.get(company_id, ""),
                company_id,
                start[:10],
                start[11:16],
                meta.get("recurrence") or "One-time",
                str(meta.get("plan_end") or ""),
                "",  # per-occurrence rows carry no weekday rule; it lives on the batch
                ", ".join(event.get("assigned_departments") or []),
                ", ".join(people.get(str(m), str(m)) for m in (event.get("assigned_member_ids") or [])),
                ", ".join(people.get(str(s), str(s)) for s in (event.get("coach_ids") or [])),
                event.get("additional_details") or "",
            ])

    rows.sort(key=lambda r: (str(r[7]), str(r[3])))   # by plan date, then title
    return rows


async def export_workbook() -> Tuple[bytes, Dict[str, int]]:
    """The whole of TPMS as one .xlsx. Returns (file bytes, {sheet name: row count})."""
    workbook = Workbook()
    counts: Dict[str, int] = {}

    # Sheet 1 — the fillable one, which is why it takes the workbook's default sheet.
    sheet = workbook.active
    sheet.title = SCHEDULE_SHEET
    sheet.append([SCHEDULE_HINT])
    sheet.append(SCHEDULE_COLUMNS)
    rows = await _schedule_rows()
    for row in rows:
        sheet.append([_cell(v) for v in row])
    counts[SCHEDULE_SHEET] = len(rows)

    for width, column in zip(
        (26, 26, 14, 30, 30, 26, 26, 14, 10, 16, 14, 16, 22, 40, 40, 40), sheet.columns
    ):
        sheet.column_dimensions[column[0].column_letter].width = width
    sheet.freeze_panes = "A3"   # keep the hint + header visible while filling rows

    # One sheet per collection, add-only on the way back in.
    for coll_name in BACKUP_COLLECTIONS:
        try:
            docs = await get_collection(coll_name).find({}).to_list(50000)
        except Exception as exc:
            logger.warning("TPMS export: could not read %s — %s", coll_name, exc)
            docs = []

        tab = workbook.create_sheet(_sheet_title(coll_name))
        # Union of keys, not the first document's: Mongo is schemaless and a later row may
        # carry a field the first one never had. `_id` is pinned first so the identity column
        # is always column A on the way back in.
        headers: List[str] = ["_id"]
        for doc in docs:
            for key in doc:
                if key not in headers:
                    headers.append(key)
        tab.append(headers)
        for doc in docs:
            tab.append([_cell(doc.get(key)) for key in headers])
        tab.freeze_panes = "A2"
        counts[coll_name] = len(docs)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), counts


# ─────────────────────────────────────────────────────────────
# Import
# ─────────────────────────────────────────────────────────────
def _read_sheet(sheet, *, header_row: int = 1) -> List[Dict[str, Any]]:
    """A worksheet as a list of {column: value}, skipping wholly blank rows."""
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < header_row:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[header_row - 1]]
    out: List[Dict[str, Any]] = []
    for raw in rows[header_row:]:
        if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        record: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if header:
                record[header] = raw[idx] if idx < len(raw) else None
        out.append(record)
    return out


async def _resolve_people(labels: List[str], collections: Tuple[str, ...]) -> Tuple[List[str], List[str]]:
    """Emails (or names, or raw ids) → Mongo ids. Returns (found ids, unresolved labels).

    Email first because that is what Export writes and what a human can copy from a mail
    client; a full name is accepted as a fallback so a hand-typed sheet still works.
    """
    found: List[str] = []
    missing: List[str] = []
    for label in labels:
        key = label.strip()
        if not key:
            continue
        hit = None
        for coll_name in collections:
            coll = get_collection(coll_name)
            if "@" in key:
                hit = await coll.find_one({"email": {"$regex": f"^{key}$", "$options": "i"}})
            if not hit:
                hit = await coll.find_one({"full_name": {"$regex": f"^{key}$", "$options": "i"}})
            if not hit:
                try:
                    hit = await coll.find_one({"_id": ObjectId(key)})
                except (InvalidId, TypeError):
                    hit = None
            if hit:
                break
        if hit:
            found.append(str(hit["_id"]))
        else:
            missing.append(key)
    return found, missing


async def _schedule_payload(row: Dict[str, Any]) -> Dict[str, Any]:
    """One sheet row → the payload `create_schedule` expects, or raise ValueError with the
    reason. Mirrors the Schedule modal's own field names so the service sees nothing new."""
    title = _text(row, "Title")
    activity = _text(row, "Activity")
    if not title:
        raise ValueError("Title is required")
    if not activity:
        raise ValueError("Activity is required")

    company_id = _text(row, "Company ID")
    company_name = _text(row, "Company Name")
    if not company_id:
        if not company_name:
            raise ValueError("Company Name (or Company ID) is required")
        company = await get_collection("companies").find_one(
            {"name": {"$regex": f"^{company_name}$", "$options": "i"}})
        if not company:
            raise ValueError(f"No company named '{company_name}'")
        company_id = str(company["_id"])
        company_name = company.get("name") or company_name

    plan_start = _date_text(row, "Plan Date")
    if not plan_start:
        raise ValueError("Plan Date is required (YYYY-MM-DD)")

    departments = _csv(row, "Departments")
    if not departments:
        raise ValueError("At least one Department is required")

    member_ids, unknown_members = await _resolve_people(
        _csv(row, "Company Assigners"), ("learners",))
    if unknown_members:
        raise ValueError(f"Unknown Company Assigner(s): {', '.join(unknown_members)}")
    if not member_ids:
        raise ValueError("At least one Company Assigner is required")

    staff_ids, unknown_staff = await _resolve_people(_csv(row, "Staff Assigners"), ("staff",))
    if unknown_staff:
        raise ValueError(f"Unknown Staff Assigner(s): {', '.join(unknown_staff)}")

    return {
        "title": title,
        "activity": activity,
        "company_id": company_id,
        "company_name": company_name,
        "plan_start": plan_start,
        "plan_end": _date_text(row, "Plan End"),
        "event_time": _time_text(row, "Time"),
        "recurrence": _text(row, "Recurrence") or "One-time",
        "weekdays": _weekdays(row),
        "departments": departments,
        "member_ids": member_ids,
        "staff_ids": staff_ids,
        "comment": _text(row, "Comment"),
    }


async def _import_schedules(sheet, user: dict) -> Dict[str, Any]:
    """Create a schedule for every row that has no Schedule ID.

    Each row is replayed through `create_schedule`, so an imported activity is indistinguishable
    from one made in the modal: recurrence expanded, reminders attached, tracker rows written,
    schedule mail sent and form links minted. Rows are independent — one bad row is reported and
    the rest still land, because refusing the whole file over a single typo would make a
    hundred-row import unusable.
    """
    # Row 1 is the hint line written by Export; the real header is row 2. A sheet that was
    # hand-made without the hint is still read correctly by falling back to row 1.
    rows = _read_sheet(sheet, header_row=2)
    if not rows or "Title" not in (rows[0] if rows else {}):
        rows = _read_sheet(sheet, header_row=1)

    created = skipped = failed = 0
    events = 0
    errors: List[str] = []

    for offset, row in enumerate(rows):
        # +3: one for the hint line, one for the header, one to make it 1-based like Excel.
        line = offset + 3
        if _text(row, "Schedule ID"):
            skipped += 1
            continue
        if not any(_text(row, col) for col in ("Title", "Activity", "Company Name", "Company ID")):
            continue
        try:
            payload = await _schedule_payload(row)
            result = await create_schedule_for_import(user, payload)
            created += 1
            events += int(result.get("count") or 0)
        except HTTPException as exc:
            failed += 1
            errors.append(f"Row {line}: {exc.detail}")
        except ValueError as exc:
            failed += 1
            errors.append(f"Row {line}: {exc}")
        except Exception as exc:
            failed += 1
            logger.error("TPMS import: row %s failed — %s", line, exc)
            errors.append(f"Row {line}: {exc}")

    return {
        "created": created,
        "events": events,
        "skipped": skipped,
        "failed": failed,
        # Capped so a workbook where every row is wrong returns a readable response rather
        # than a wall of text; the count above still reports the true total.
        "errors": errors[:50],
    }


async def create_schedule_for_import(user: dict, payload: dict) -> dict:
    """Thin seam over `create_schedule` — imported here rather than at module load because
    tpms_schedule_service imports notification code that pulls in this module's siblings."""
    from app.services.tpms_schedule_service import create_schedule

    return await create_schedule(user, payload)


async def _import_collection(sheet, coll_name: str) -> Dict[str, int]:
    """Add-only restore of one collection sheet: insert rows whose `_id` is not already there."""
    rows = _read_sheet(sheet)
    if not rows:
        return {"inserted": 0, "skipped": 0}

    coll = get_collection(coll_name)
    inserted = skipped = 0

    for row in rows:
        raw_id = row.get("_id")
        doc: Dict[str, Any] = {}
        for key, cell in row.items():
            if key == "_id":
                continue
            parsed = _value(cell)
            if parsed is not None:
                doc[key] = parsed
        if not doc:
            continue

        if raw_id is not None and str(raw_id).strip():
            try:
                oid: Any = ObjectId(str(raw_id).strip())
            except (InvalidId, TypeError):
                oid = str(raw_id).strip()
            if await coll.find_one({"_id": oid}):
                skipped += 1
                continue
            doc["_id"] = oid

        await coll.insert_one(doc)
        inserted += 1

    return {"inserted": inserted, "skipped": skipped}


async def import_workbook(content: bytes, user: dict) -> Dict[str, Any]:
    """Load an exported workbook back in. See the module docstring for the two import paths."""
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Not a readable .xlsx workbook: {exc}")

    report: Dict[str, Any] = {"ok": True, "collections": {}}

    if SCHEDULE_SHEET in workbook.sheetnames:
        report["schedules"] = await _import_schedules(workbook[SCHEDULE_SHEET], user)
    else:
        report["schedules"] = {
            "created": 0, "events": 0, "skipped": 0, "failed": 0,
            "errors": [f"No '{SCHEDULE_SHEET}' sheet in this workbook."],
        }

    for coll_name in BACKUP_COLLECTIONS:
        title = _sheet_title(coll_name)
        if title not in workbook.sheetnames:
            continue
        try:
            report["collections"][coll_name] = await _import_collection(workbook[title], coll_name)
        except Exception as exc:
            logger.error("TPMS import: sheet %s failed — %s", title, exc)
            report["collections"][coll_name] = {"inserted": 0, "skipped": 0, "error": str(exc)}

    report["total_inserted"] = (
        report["schedules"]["created"]
        + sum(c.get("inserted", 0) for c in report["collections"].values())
    )
    return report
