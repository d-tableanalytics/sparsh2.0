"""HRMS > recruitment analytics and reports. READ-ONLY.

Closes the gap both analysis documents flag: the source computed every figure in the
browser (FRONTEND_ANALYSIS 6.1, BACKEND_ANALYSIS "no reporting backend"). That approach has
three problems this module does not:

  1. It cannot be role-scoped. Whatever the browser is sent, it can total -- so a hiring
     manager who should see their own requisitions is one devtools panel away from the
     company's whole pipeline. Here every aggregation runs behind `_scope`.
  2. It does not scale. Totalling 10k candidates client-side means shipping 10k candidates.
  3. It drifts. Each screen re-derives "how many were interviewed" slightly differently.
     Here the funnel is declared once, in FUNNEL_STAGES, and computed once.

-- Effective rank ------------------------------------------------------------------
A funnel that counts `application_status` can show more offers than interviews, because a
candidate at `Offer Accepted` no longer has an interview status. Every candidate is
therefore ranked by the furthest point they can be SHOWN to have reached -- their status,
OR the existence of an assessment, interview or offer record, whichever is furthest. The
funnel then counts "reached at least this stage", which is monotonically non-increasing by
construction. See models/hrms.py STAGE_RANK.

-- Nothing here writes ---------------------------------------------------------------
No insert, no update, no delete. Analytics that mutate are a debugging nightmare, and a
read-only surface can be reasoned about from its inputs alone.

-- Exports are generated server-side ---------------------------------------------------
Deliberately not built in the browser from a fetched page. The rows are already
role-scoped and paginated here; rebuilding a file client-side would mean shipping rows the
UI had correctly withheld. This mirrors routes/reports.py, but does NOT import from it --
that module is outside HRMS's scope and coupling to its private helpers would make this
module break when it changes.
"""
import csv
import io
import re
from datetime import datetime, timedelta, timezone
from typing import Optional

from bson import ObjectId
from bson.errors import InvalidId
from fastapi import HTTPException

from app.db.mongodb import get_collection
from app.models.hrms import (
    BREAKDOWN_FIELDS, COLL_ASSESSMENTS, COLL_CANDIDATES, COLL_INTERVIEWS,
    COLL_JOB_POSTINGS, COLL_OFFERS, COLL_ONBOARDING, COLL_REQUISITIONS,
    FUNNEL_STAGES, MAX_BREAKDOWN_ROWS, MAX_EXPORT_ROWS, MAX_RANGE_DAYS,
    MAX_REPORT_PAGE_SIZE, RANK_IF_ACCEPTED, RANK_IF_ASSESSED, RANK_IF_INTERVIEWED,
    RANK_IF_OFFERED, REPORT_ENTITIES, SALARY_REPORT_COLUMNS, AppStatus, Cap, HrmsRole,
    InterviewStatus, OfferStatus, OnboardStatus, ReqApproval, ReqClosing, conversion,
    is_iso_date, stage_rank,
)
# ── Phase 11-R (Items 4, 6) ──
from app.models.hrms import ClientShareStatus, budget_status
# ── Internal recruitment track ──
from app.models.hrms import (
    COLL_EMPLOYEE_PROFILES, COLL_PROBATION_REVIEWS, COLL_REFERENCE_CHECKS,
    REFERENCE_CLEARS_OFFER, ProbationOutcome, RequisitionTrack,
)
from app.utils.hrms_access import can, hrms_role

# Every read in this module is capped. An unbounded to_list on an analytics endpoint is a
# denial-of-service waiting for the first client with real volume.
SCAN_CAP = 20000


# ─────────────────────────────────────────────────────────────
# Scoping — the security boundary of this whole module
# ─────────────────────────────────────────────────────────────
async def _manager_requisitions(actor: dict, company_id: str) -> list:
    """The requisition numbers a hiring manager owns.

    Mirrors hrms_candidate_service._scope_filter exactly. Duplicating the rule would let
    the two drift, and the one that drifts is the one that leaks -- so this reads the same
    source of truth (requisitions raised by this actor) and fails CLOSED: an empty list
    means an `$in: []`, which matches nothing.
    """
    rows = await get_collection(COLL_REQUISITIONS).find(
        {"company_id": str(company_id), "created_by": str(actor.get("_id"))},
        {"request_no": 1}).to_list(2000)
    return [r["request_no"] for r in rows if r.get("request_no")]


async def _client_requisitions(company_id: str, client_id: str) -> list:
    """The requisition numbers belonging to one client (Phase 11-R, Item 4).

    Fails CLOSED in the same way `_manager_requisitions` does: a client with no
    requisitions yields `$in: []`, which matches nothing rather than everything. A filter
    that silently widens when it finds nothing is how a scoping bug becomes a data leak.
    """
    rows = await get_collection(COLL_REQUISITIONS).find(
        {"company_id": str(company_id), "client_id": str(client_id)},
        {"request_no": 1}).to_list(5000)
    return [r["request_no"] for r in rows if r.get("request_no")]


async def _scope(actor: dict, company_id: str, client_id: str = None) -> dict:
    """The base `$match` for every aggregation in this module.

    A MANAGER is narrowed to their own requisitions. Everyone else with `analytics.read`
    sees the company. This is applied to EVERY query below without exception -- there is no
    "just this one summary" path that skips it.

    Phase 11-R adds an optional CLIENT narrowing. It composes with the manager narrowing by
    INTERSECTION, never replacing it: a hiring manager filtering by client sees the
    requisitions that are both theirs and that client's, so the filter can only ever narrow
    what they may see, never widen it. `company_id` remains the only tenant boundary --
    `client_id` is a reporting dimension inside one tenant, not a second security rule.
    """
    base = {"company_id": str(company_id)}
    is_manager = hrms_role(actor) == HrmsRole.MANAGER

    if not client_id:
        if not is_manager:
            return base
        return {**base,
                "request_no": {"$in": await _manager_requisitions(actor, company_id)}}

    client_reqs = await _client_requisitions(company_id, client_id)
    if not is_manager:
        return {**base, "request_no": {"$in": client_reqs}}

    own = set(await _manager_requisitions(actor, company_id))
    return {**base, "request_no": {"$in": [r for r in client_reqs if r in own]}}


def _scoped_by_request(scope: dict, field: str = "request_no") -> dict:
    """Re-express a scope for a collection whose link to a requisition uses another name.

    Every HRMS collection carries `request_no`, so in practice this is the identity -- it
    exists so a future collection that does not can be handled in one place rather than by
    quietly dropping the manager narrowing.
    """
    out = dict(scope)
    if field != "request_no" and "request_no" in out:
        out[field] = out.pop("request_no")
    return out


# ─────────────────────────────────────────────────────────────
# Date range
# ─────────────────────────────────────────────────────────────
def parse_range(date_from: Optional[str], date_to: Optional[str]) -> tuple:
    """Validate and normalise a date window. Returns (start, end) as aware datetimes.

    Defaults to the last 90 days -- a dashboard with no window is a dashboard that gets
    slower every month it is in production.
    """
    now = datetime.now(timezone.utc)
    for value, label in ((date_from, "Start date"), (date_to, "End date")):
        if value and not is_iso_date(value):
            raise HTTPException(
                status_code=422,
                detail=f"{label} must be a valid date in YYYY-MM-DD format.")

    end = (datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=timezone.utc)
           if date_to else now)
    start = (datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=timezone.utc)
             if date_from else end - timedelta(days=90))

    # End of the chosen day, so "to = today" includes everything that happened today.
    end = end.replace(hour=23, minute=59, second=59, microsecond=999999)

    if start > end:
        raise HTTPException(
            status_code=422, detail="The start date must be on or before the end date.")
    if (end - start).days > MAX_RANGE_DAYS:
        raise HTTPException(
            status_code=422,
            detail=f"Choose a range of {MAX_RANGE_DAYS} days or fewer.")
    return start, end


def _window(field: str, start: datetime, end: datetime) -> dict:
    return {field: {"$gte": start, "$lte": end}}


# ─────────────────────────────────────────────────────────────
# Effective rank
# ─────────────────────────────────────────────────────────────
async def _evidence_ranks(scope: dict) -> dict:
    """`uk` -> the rank implied by records elsewhere in the pipeline.

    Four cheap grouped reads rather than a `$lookup` chain: each is index-backed on its own
    collection, the result sets are small (one row per candidate touched), and the merge is
    plain arithmetic that a person can check by hand.
    """
    ranks = {}

    def bump(uk, rank):
        if uk and ranks.get(uk, 0) < rank:
            ranks[uk] = rank

    for coll, floor in ((COLL_ASSESSMENTS, RANK_IF_ASSESSED),
                        (COLL_INTERVIEWS, RANK_IF_INTERVIEWED)):
        rows = await get_collection(coll).find(scope, {"uk": 1}).to_list(SCAN_CAP)
        for row in rows:
            bump(row.get("uk"), floor)

    # An offer proves stage 6; an ACCEPTED offer proves stage 7. A draft proves neither --
    # it has not been issued, so nothing has happened to the candidate yet.
    offers = await get_collection(COLL_OFFERS).find(
        scope, {"uk": 1, "status": 1}).to_list(SCAN_CAP)
    for row in offers:
        status = row.get("status")
        if status == OfferStatus.DRAFT.value:
            continue
        bump(row.get("uk"),
             RANK_IF_ACCEPTED if status == OfferStatus.ACCEPTED.value else RANK_IF_OFFERED)

    return ranks


def _effective(candidate: dict, evidence: dict) -> int:
    return max(stage_rank(candidate.get("application_status")),
               evidence.get(candidate.get("uk"), 0))


# ─────────────────────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────────────────────
async def dashboard(actor: dict, company_id: str, *, date_from: str = None,
                    date_to: str = None, client_id: str = None,
                    track: str = None) -> dict:
    """Headline KPIs plus the positions/vacancy summary.

    Every tile carries the `link` and `filter` the UI needs to deep-link into the screen
    that produced the number. A KPI you cannot click through to is a number the reader has
    to take on trust.
    """
    start, end = parse_range(date_from, date_to)
    scope = await _scope(actor, company_id, client_id)

    candidates = await get_collection(COLL_CANDIDATES).find(
        {**scope, **_window("applied_at", start, end)},
        # Phase 11-R: client_share and the referral fields ride along, so the new tiles are
        # computed from the SAME single read rather than adding a second pass.
        {"uk": 1, "application_status": 1, "source": 1, "applied_at": 1,
         "client_share": 1, "client_share_status": 1, "is_referral": 1,
         "referral_source": 1}).to_list(SCAN_CAP)
    evidence = await _evidence_ranks(scope)

    reqs = await get_collection(COLL_REQUISITIONS).find(
        {**scope, **_window("created_at", start, end)},
        {"request_no": 1, "vacancy": 1, "closing_status": 1,
         "approval_status": 1}).to_list(SCAN_CAP)

    interviews = await get_collection(COLL_INTERVIEWS).find(
        {**scope, **_window("scheduled_at", start, end)},
        {"status": 1, "outcome": 1, "scheduled_at": 1}).to_list(SCAN_CAP)

    offers = await get_collection(COLL_OFFERS).find(
        {**scope, **_window("created_at", start, end)},
        {"status": 1, "uk": 1, "responded_at": 1}).to_list(SCAN_CAP)

    onboardings = await get_collection(COLL_ONBOARDING).find(
        {**scope, **_window("created_at", start, end)},
        {"status": 1, "employee_id": 1}).to_list(SCAN_CAP)

    ranks = [_effective(c, evidence) for c in candidates]
    hired = sum(1 for r in ranks if r >= 7)
    in_pipeline = sum(1 for c, r in zip(candidates, ranks)
                      if c.get("application_status") not in
                      (AppStatus.REJECTED.value, AppStatus.DUPLICATE.value,
                       AppStatus.OFFER_DECLINED.value, AppStatus.EMPLOYEE_CREATED.value))

    open_reqs = [r for r in reqs if r.get("closing_status") == ReqClosing.OPEN.value]
    total_vacancy = sum(int(r.get("vacancy") or 1) for r in open_reqs)
    awaiting = sum(1 for r in reqs if r.get("approval_status") in
                   (ReqApproval.PENDING_HR.value, ReqApproval.PENDING_MD.value))

    offers_sent = sum(1 for o in offers if o.get("status") != OfferStatus.DRAFT.value)
    offers_accepted = sum(1 for o in offers if o.get("status") == OfferStatus.ACCEPTED.value)

    ttf = await _time_to_hire(scope, start, end)
    cv = _cv_metrics(candidates, ranks)

    return {
        "range": {"from": start.strftime("%Y-%m-%d"), "to": end.strftime("%Y-%m-%d")},
        "scoped_to_own_requisitions": hrms_role(actor) == HrmsRole.MANAGER,
        # ── Phase 11-R, Item 4 ──
        "client_id": client_id,
        "cv_metrics": cv,
        # CV review -> selection -> client sharing -> client verdict -> joining, as stages.
        "cv_funnel": _cv_funnel(cv),
        # ── Internal track ── the SOP §10 KPI block, present only when asked for. Omitting
        # it by default keeps the existing payload byte-for-byte for every caller written
        # before the internal track existed.
        "internal_kpis": (await internal_kpis(actor, company_id, date_from=date_from,
                                              date_to=date_to)
                          if track == RequisitionTrack.INTERNAL.value else None),
        "client_metrics": {
            "shared": cv["shared_with_client"],
            "shortlisted": cv["client_shortlisted"],
            "rejected": cv["client_rejected"],
            "awaiting_verdict": cv["client_awaiting"],
            # Of the CVs the client has actually ANSWERED on. Measuring against everything
            # shared would report a low rate for a client who is simply slow, which is a
            # different problem and would read as a quality one.
            "shortlist_rate": conversion(
                cv["client_shortlisted"],
                cv["client_shortlisted"] + cv["client_rejected"]),
        },
        # Populated only when no single client is selected — the comparison IS the
        # "all clients" view.
        "client_comparison": (None if client_id else
                              await _client_comparison(actor, company_id, start, end)),
        "kpis": [
            {"key": "candidates", "label": "Candidates", "value": len(candidates),
             "hint": "Applications received in this period",
             "link": "/hrms/candidates"},
            {"key": "in_pipeline", "label": "In pipeline", "value": in_pipeline,
             "hint": "Still live -- not rejected, declined or already hired",
             "link": "/hrms/candidates"},
            {"key": "open_requisitions", "label": "Open requisitions",
             "value": len(open_reqs),
             "hint": f"{total_vacancy} vacancies to fill",
             "link": "/hrms/requisitions"},
            {"key": "awaiting_approval", "label": "Awaiting approval", "value": awaiting,
             "hint": "Requisitions with HR or MD",
             "link": "/hrms/requisitions"},
            {"key": "interviews", "label": "Interviews", "value": len(interviews),
             "hint": f"{sum(1 for i in interviews if i.get('status') == InterviewStatus.COMPLETED.value)} completed",
             "link": "/hrms/interviews"},
            {"key": "offers_sent", "label": "Offers sent", "value": offers_sent,
             "hint": f"{offers_accepted} accepted",
             "link": "/hrms/offers"},
            {"key": "onboarding", "label": "Onboarding", "value": len(onboardings),
             "hint": f"{sum(1 for o in onboardings if o.get('employee_id'))} employee IDs issued",
             "link": "/hrms/onboarding"},
            {"key": "hired", "label": "Hired", "value": hired,
             "hint": "Reached an accepted offer or beyond",
             "link": "/hrms/onboarding"},
            # ── Phase 11-R, Item 4 ── every new tile carries the same link + filter the
            # existing ones do, so a reader can click through to the rows behind the number.
            {"key": "cvs_reviewed", "label": "CVs reviewed",
             "value": cv["reviewed"],
             "hint": f"{cv['awaiting_review']} still awaiting review",
             "link": "/hrms/candidates"},
            {"key": "cvs_shortlisted", "label": "CVs shortlisted",
             "value": cv["shortlisted"],
             "hint": "Cleared our own screening",
             "link": "/hrms/candidates",
             "filter": {"status": AppStatus.SHORTLISTED.value}},
            {"key": "cvs_selected", "label": "CVs selected", "value": cv["selected"],
             "hint": "Reached Selected or further",
             "link": "/hrms/candidates", "filter": {"status": AppStatus.SELECTED.value}},
            {"key": "cvs_rejected", "label": "CVs rejected", "value": cv["rejected"],
             "hint": "Rejected, declined, duplicated or failed",
             "link": "/hrms/candidates", "filter": {"status": AppStatus.REJECTED.value}},
            {"key": "shared_with_client", "label": "Shared with client",
             "value": cv["shared_with_client"],
             "hint": f"{cv['client_awaiting']} awaiting a verdict",
             "link": "/hrms/candidates",
             "filter": {"status": AppStatus.SHARED_WITH_CLIENT.value}},
            {"key": "client_shortlisted", "label": "Client shortlisted",
             "value": cv["client_shortlisted"],
             "hint": "The client's own shortlist",
             "link": "/hrms/candidates",
             "filter": {"status": AppStatus.CLIENT_SHORTLISTED.value}},
            {"key": "client_rejected", "label": "Client rejections",
             "value": cv["client_rejected"],
             "hint": "Rejected by the client after review",
             "link": "/hrms/candidates",
             "filter": {"status": AppStatus.CLIENT_REJECTED.value}},
            {"key": "joinings", "label": "Total joinings", "value": cv["joinings"],
             "hint": "Joined or converted to an employee record",
             "link": "/hrms/onboarding"},
        ],
        "positions": {
            "open": len(open_reqs),
            "vacancies": total_vacancy,
            "filled": sum(1 for r in reqs if r.get("closing_status") == ReqClosing.HIRED.value),
            "on_hold": sum(1 for r in reqs if r.get("closing_status") == ReqClosing.HOLD.value),
            "cancelled": sum(1 for r in reqs if r.get("closing_status") == ReqClosing.CANCEL.value),
        },
        "offer_outcomes": {
            "draft": sum(1 for o in offers if o.get("status") == OfferStatus.DRAFT.value),
            "sent": sum(1 for o in offers if o.get("status") == OfferStatus.SENT.value),
            "accepted": offers_accepted,
            "declined": sum(1 for o in offers if o.get("status") == OfferStatus.DECLINED.value),
            "revoked": sum(1 for o in offers if o.get("status") == OfferStatus.REVOKED.value),
            "acceptance_rate": conversion(
                offers_accepted,
                sum(1 for o in offers if o.get("status") in
                    (OfferStatus.ACCEPTED.value, OfferStatus.DECLINED.value))),
        },
        "onboarding_states": {
            "pre_onboarding": sum(1 for o in onboardings
                                  if o.get("status") == OnboardStatus.PRE_ONBOARDING.value),
            "onboarding": sum(1 for o in onboardings
                              if o.get("status") == OnboardStatus.ONBOARDING.value),
            "completed": sum(1 for o in onboardings
                             if o.get("status") == OnboardStatus.COMPLETED.value),
        },
        "time_to_hire": ttf,
    }


# ─────────────────────────────────────────────────────────────
# Phase 11-R, Item 4 — CV and client metrics
# ─────────────────────────────────────────────────────────────
# Every figure below is DERIVED from data that already exists after the client-share write
# path lands, and every one is computed from the ALREADY-SCOPED candidate list rather than
# by issuing its own query — so none of them can accidentally escape `_scope`.
#
# The formulas, stated once here and repeated in PHASE_11R_REPORT:
#
#   reviewed    = somebody has ACTED on the CV: it cleared the shortlist bar, or it carries
#                 a status that is itself a review outcome (Under Review, Rejected,
#                 Duplicate, On Hold). A CV sitting at Applied has not been reviewed.
#   shortlisted = effective_rank >= rank(Shortlisted)  -- the internal selection that
#                 precedes sharing anything with a client
#   selected    = effective_rank >= rank(Selected)     -- the FINAL selection, after
#                 assessment and interview
#   rejected    = status in the rejection set (a STATUS, not a rank: a rejection is a
#                 destination, and ranking it where they entered is what keeps the funnel
#                 monotonic — see STAGE_RANK)
#   joinings    = status in {Joined, Employee Created}
#
# Everything rank-based uses EFFECTIVE rank, so a candidate now at Offer Accepted still
# counts as reviewed and shortlisted — exactly the bug the Phase 10 rank system was built to
# prevent.
#
# On `reviewed` specifically: it cannot be `rank >= rank(Under Review)`, because STAGE_RANK
# puts Applied and Under Review in the SAME band (both rank 1, deliberately — neither has
# cleared a hiring gate). That test therefore counted every candidate with a recognised
# status, and "CVs reviewed" silently equalled "CVs received" on every dashboard that showed
# both. The set below is the smallest honest correction: rank 1 statuses that DO represent a
# human decision, plus everything above the band.
REJECTION_STATUSES = {
    AppStatus.REJECTED.value, AppStatus.CLIENT_REJECTED.value, AppStatus.DUPLICATE.value,
    AppStatus.ASSESSMENT_FAILED.value, AppStatus.OFFER_DECLINED.value,
}
JOINED_STATUSES = {AppStatus.JOINED.value, AppStatus.EMPLOYEE_CREATED.value}
# Rank-1 statuses that mean the CV was looked at, as opposed to merely received.
REVIEW_OUTCOME_STATUSES = {
    AppStatus.UNDER_REVIEW.value, AppStatus.REJECTED.value, AppStatus.DUPLICATE.value,
    AppStatus.ON_HOLD.value,
}


def _cv_metrics(candidates: list, ranks: list) -> dict:
    """The Item 4 headline figures, from one already-scoped pass over the candidates."""
    rank_shortlisted = stage_rank(AppStatus.SHORTLISTED.value)
    rank_selected = stage_rank(AppStatus.SELECTED.value)

    shared = shortlisted = rejected_by_client = awaiting = 0
    for c in candidates:
        share = c.get("client_share") or {}
        # `shared_at` is the fact of the share; `client_share_status` is the denormalised
        # verdict. Reading the sub-document first means a row written before the flat field
        # existed still counts.
        if not share.get("shared_at"):
            continue
        shared += 1
        verdict = share.get("status") or c.get("client_share_status")
        if verdict == ClientShareStatus.SHORTLISTED.value:
            shortlisted += 1
        elif verdict == ClientShareStatus.REJECTED.value:
            rejected_by_client += 1
        elif verdict in (None, ClientShareStatus.PENDING.value):
            awaiting += 1

    reviewed = sum(1 for c, r in zip(candidates, ranks)
                   if r >= rank_shortlisted
                   or c.get("application_status") in REVIEW_OUTCOME_STATUSES)

    return {
        "reviewed": reviewed,
        # What is still sitting in the inbox. Stated explicitly so "received minus reviewed"
        # is a figure the dashboard shows rather than one the reader has to work out.
        "awaiting_review": len(candidates) - reviewed,
        "shortlisted": sum(1 for r in ranks if r >= rank_shortlisted),
        "selected": sum(1 for r in ranks if r >= rank_selected),
        "rejected": sum(1 for c in candidates
                        if c.get("application_status") in REJECTION_STATUSES),
        "shared_with_client": shared,
        "client_shortlisted": shortlisted,
        "client_rejected": rejected_by_client,
        "client_awaiting": awaiting,
        "joinings": sum(1 for c in candidates
                        if c.get("application_status") in JOINED_STATUSES),
        "referrals": sum(1 for c in candidates if c.get("is_referral")),
        "total": len(candidates),
    }


# The recruitment funnel this dashboard exists to show, declared ONCE:
#
#   CV received -> reviewed -> shortlisted -> shared with client -> client shortlisted
#               -> selected -> joined
#
# Each entry is (key, label, hint). The values come straight from `_cv_metrics`, so the
# funnel and the KPI tiles cannot disagree -- they are the same numbers, arranged.
CV_FUNNEL_STAGES = [
    ("total",              "CVs received",       "Applications in this period"),
    ("reviewed",           "Reviewed",           "Somebody has acted on the CV"),
    ("shortlisted",        "Shortlisted",        "Cleared our own screening"),
    ("shared_with_client", "Shared with client", "Sent to the client for their verdict"),
    ("client_shortlisted", "Client shortlisted", "The client's own shortlist"),
    ("selected",           "Selected",           "Final selection after interview"),
    ("joinings",           "Joined",             "Joined or converted to an employee"),
]


def _cv_funnel(cv: dict) -> list:
    """The CV funnel as stages, with the drop-off between each pair.

    `of_previous` is null rather than a number over 100% wherever a stage exceeds the one
    above it. That is not defensive rounding -- it is a real and MEANINGFUL shape: an
    in-house requisition never shares a CV with a client, so `selected` can legitimately
    exceed `shared_with_client`. Printing "340%" there would read as a bug in the dashboard
    rather than what it is, a mix of client and in-house hiring in one view.

    `of_total` is always against CVs received, so every bar stays comparable.
    """
    total = cv.get("total") or 0
    out = []
    previous = None
    for key, label, hint in CV_FUNNEL_STAGES:
        value = cv.get(key) or 0
        out.append({
            "key": key,
            "label": label,
            "hint": hint,
            "value": value,
            "of_total": conversion(value, total),
            "of_previous": (None if previous is None or value > previous
                            else conversion(value, previous)),
        })
        previous = value
    return out


# ─────────────────────────────────────────────────────────────
# Internal track — the KPI block (SOP §10)
# ─────────────────────────────────────────────────────────────
# Eight figures, each with the SOP's own target beside it.
#
# Every one is a RATIO, and every ratio reports its numerator and denominator alongside the
# percentage. "83%" with no denominator is unreadable: it could be five of six or five
# hundred of six hundred, and those call for different conversations.
#
# Where the denominator is zero the value is null with a reason, never 0% and never 100%.
# Both of those are claims about performance; "nothing has happened yet" is not.
# -- Denominator honesty -------------------------------------------------------------------
# Every ratio reports `eligible_n` beside the percentage, and `eligible_n` is the number of
# records that COULD have met the rule -- not the number that exist.
#
# The 90-day retention KPI is where this matters most and is easiest to get wrong. Somebody
# who joined last week has not failed to stay 90 days; they have not had 90 days. Counting
# them in the denominator scores the company down for hiring recently, and counting them as
# retained scores it up for the same thing. Both are wrong; excluding them is the only
# honest option, and saying how many were excluded is what makes the figure readable.
def _ratio(key: str, label: str, hits: int, total: int, target,
           hint: str = None, reason: str = None, excluded: int = 0,
           excluded_reason: str = None) -> dict:
    value = None if not total else round((hits / total) * 100, 1)
    return {
        "key": key, "label": label,
        "value": value, "target": target,
        "numerator": hits, "denominator": total,
        # The same number as `denominator`, named for what it MEANS rather than for where it
        # sits in the arithmetic. Callers render this one; the pair above is the workings.
        "eligible_n": total,
        # Records deliberately left out, and why. Zero on most KPIs; non-zero is the honest
        # answer to "why is this only counting eleven of our forty joiners".
        "excluded_n": excluded,
        "excluded_reason": excluded_reason,
        "meets_target": (None if value is None or target is None
                         else value >= target),
        "hint": hint,
        # Present only when there is no figure, so a caller can say WHY rather than
        # rendering a dash nobody can interpret.
        "reason": reason or (None if total else "No qualifying records in this period yet."),
    }


async def internal_kpis(actor: dict, company_id: str, *, date_from: str = None,
                        date_to: str = None) -> dict:
    """The internal recruitment KPI dashboard (SOP §10).

    Read-only, like everything else here, and behind the same `_scope`, the same SCAN_CAP and
    the same window validation. Nothing in this function is derived in the browser.
    """
    from app.services.hrms_sla_service import working_days_between

    start, end = parse_range(date_from, date_to)
    scope = await _scope(actor, company_id)

    reqs = await get_collection(COLL_REQUISITIONS).find(
        {**scope, "requisition_track": RequisitionTrack.INTERNAL.value},
        {"request_no": 1, "created_at": 1, "sla_actuals": 1,
         "approval_status": 1, "designation_name": 1}).to_list(SCAN_CAP)
    request_nos = [r["request_no"] for r in reqs if r.get("request_no")]
    if not request_nos:
        return {"applicable": False,
                "reason": "No internal requisitions have been raised yet.",
                "range": {"from": start.strftime("%Y-%m-%d"),
                          "to": end.strftime("%Y-%m-%d")},
                "kpis": []}

    scoped = {"company_id": str(company_id), "request_no": {"$in": request_nos}}

    candidates = await get_collection(COLL_CANDIDATES).find(
        scoped, {"uk": 1, "request_no": 1, "application_status": 1,
                 "applied_at": 1}).to_list(SCAN_CAP)
    offers = await get_collection(COLL_OFFERS).find(
        scoped, {"offer_no": 1, "uk": 1, "request_no": 1, "status": 1, "sent_at": 1,
                 "created_at": 1, "joining_date": 1}).to_list(SCAN_CAP)
    references = await get_collection(COLL_REFERENCE_CHECKS).find(
        scoped, {"uk": 1, "outcome": 1, "created_at": 1}).to_list(SCAN_CAP)
    probations = await get_collection(COLL_PROBATION_REVIEWS).find(
        scoped, {"prb_no": 1, "employee_code": 1, "outcome": 1, "ends_on": 1,
                 "confirmed_at": 1}).to_list(SCAN_CAP)
    onboardings = await get_collection(COLL_ONBOARDING).find(
        scoped, {"onb_no": 1, "employee_id": 1, "joining_date": 1,
                 "checklist": 1}).to_list(SCAN_CAP)

    kpis = []

    # ── 1. Budget approved before sourcing (target 100%) ──
    # Measured against requisitions that actually SOURCED. One with no candidates has not
    # been tested against the rule, and counting it as a pass would flatter the figure.
    first_cv = {}
    for c in candidates:
        key, applied = c.get("request_no"), c.get("applied_at")
        if key and applied and (key not in first_cv or applied < first_cv[key]):
            first_cv[key] = applied
    sourced = [r for r in reqs if r.get("request_no") in first_cv]
    compliant = 0
    for r in sourced:
        approved = (r.get("sla_actuals") or {}).get("budget_approved")
        if approved and approved <= first_cv[r["request_no"]]:
            compliant += 1
    kpis.append(_ratio(
        "budget_before_sourcing", "Budget approved before sourcing",
        compliant, len(sourced), 100,
        hint="Of internal requisitions that received a CV."))

    # ── 2. Shortlist ready within 15 working days (target 95%) ──
    shortlisted_reqs = [r for r in reqs
                        if (r.get("sla_actuals") or {}).get("shortlist_ready")]
    within = 0
    for r in shortlisted_reqs:
        taken = working_days_between(r.get("created_at"),
                                     r["sla_actuals"]["shortlist_ready"])
        if taken is not None and taken <= 15:
            within += 1
    kpis.append(_ratio(
        "shortlist_within_tat", "Shortlist ready within Day 15",
        within, len(shortlisted_reqs), 95,
        hint="Working days from the requisition being raised."))

    # ── 3. Offer-to-joining conversion (tracked, no fixed target) ──
    joined_uks = {c["uk"] for c in candidates
                  if c.get("application_status") in JOINED_STATUSES}
    sent_offers = [o for o in offers if o.get("status") != OfferStatus.DRAFT.value]
    converted = sum(1 for o in sent_offers if o.get("uk") in joined_uks)
    kpis.append(_ratio(
        "offer_to_joining", "Offer-to-joining conversion",
        converted, len(sent_offers), None,
        hint="Of offers that went out, how many resulted in somebody joining."))

    # ── 4. Reference check completed before offer (target 100%) ──
    # "Before" is compared against the offer's CREATION, because that is the moment the gate
    # runs. Comparing against the send would call a reference taken after the letter was
    # drafted compliant.
    clearing_by_uk = {}
    for ref in references:
        if ref.get("outcome") not in REFERENCE_CLEARS_OFFER:
            continue
        uk, at = ref.get("uk"), ref.get("created_at")
        if uk and at and (uk not in clearing_by_uk or at < clearing_by_uk[uk]):
            clearing_by_uk[uk] = at
    referenced = 0
    for o in offers:
        cleared_at = clearing_by_uk.get(o.get("uk"))
        if cleared_at and o.get("created_at") and cleared_at <= o["created_at"]:
            referenced += 1
    kpis.append(_ratio(
        "reference_before_offer", "Reference check before offer",
        referenced, len(offers), 100,
        hint="Anything short of 100% was let through on an approved exception."))

    # ── 5. Offer letter issued before the joining date (target 100%) ──
    issued_in_time = 0
    datable = [o for o in sent_offers if o.get("sent_at") and o.get("joining_date")]
    for o in datable:
        sent_on = o["sent_at"].strftime("%Y-%m-%d") if hasattr(o["sent_at"], "strftime") \
            else str(o["sent_at"])[:10]
        if sent_on <= str(o["joining_date"])[:10]:
            issued_in_time += 1
    kpis.append(_ratio(
        "offer_before_joining", "Offer letter issued before joining",
        issued_in_time, len(datable), 100))

    # ── 6. Probation confirmations completed on time (target 95%) ──
    decided = [p for p in probations
               if p.get("outcome") != ProbationOutcome.PENDING.value]
    on_time = 0
    for p in decided:
        confirmed = p.get("confirmed_at")
        if not confirmed or not p.get("ends_on"):
            continue
        decided_on = confirmed.strftime("%Y-%m-%d") if hasattr(confirmed, "strftime") \
            else str(confirmed)[:10]
        if decided_on <= str(p["ends_on"])[:10]:
            on_time += 1
    kpis.append(_ratio(
        "probation_on_time", "Probation confirmed on time",
        on_time, len(decided), 95,
        hint="Decided on or before the probation end date."))

    # ── 7. 90-day retention of new joinees ──
    # Only joiners who have HAD 90 days count. Including somebody who joined last week would
    # score them as retained for a period they have not lived through.
    cutoff = datetime.now(timezone.utc) - timedelta(days=90)
    codes = [o["employee_id"] for o in onboardings if o.get("employee_id")]
    eligible = retained = 0
    immature = undated = 0
    if codes:
        profiles = await get_collection(COLL_EMPLOYEE_PROFILES).find(
            {"company_id": str(company_id), "employee_code": {"$in": codes}},
            # `separation_date` is Phase INT-2's addition and `resigned_on` is the field
            # that was already there. Both are read, newest-named first, so a company that
            # has been recording one of them keeps its figure -- see `_separation_date`.
            {"employee_code": 1, "joined_on": 1, "resigned_on": 1, "separation_date": 1,
             "employment_status": 1, "status": 1}).to_list(SCAN_CAP)
        for p in profiles:
            joined = p.get("joined_on")
            if not joined:
                undated += 1
                continue
            try:
                joined_at = datetime.strptime(str(joined)[:10], "%Y-%m-%d").replace(
                    tzinfo=timezone.utc)
            except ValueError:
                undated += 1
                continue
            if joined_at > cutoff:
                # THE DENOMINATOR-HONESTY CASE. Their 90-day window has not matured, so they
                # are neither retained nor lost -- they are not yet measurable. Counted as
                # excluded and reported, never folded into either side of the ratio.
                immature += 1
                continue
            eligible += 1
            left = _separation_date(p)
            if not left:
                retained += 1
                continue
            try:
                left_at = datetime.strptime(str(left)[:10], "%Y-%m-%d").replace(
                    tzinfo=timezone.utc)
            except ValueError:
                # An unreadable date is not evidence somebody left inside 90 days. Counting
                # it against them would let a typo depress a retention figure.
                retained += 1
                continue
            if (left_at - joined_at).days >= 90:
                retained += 1
    excluded = immature + undated
    excluded_bits = []
    if immature:
        excluded_bits.append(f"{immature} joiner(s) have not completed 90 days yet")
    if undated:
        excluded_bits.append(f"{undated} have no joining date recorded")
    kpis.append(_ratio(
        "retention_90_day", "90-day retention of new joinees",
        retained, eligible, None,
        hint="Counts only joiners whose 90-day window has matured.",
        excluded=excluded,
        excluded_reason="; ".join(excluded_bits) or None,
        reason=(None if eligible else
                "No internal joiner has completed 90 days yet.")))

    # ── 8. New-hire satisfaction ──
    # Phase INT-2 made this measurable. Until then the induction checklist recorded WHETHER
    # feedback was collected and the SOP asked for a SCORE, and averaging ticks into a
    # number would have invented data -- so the KPI honestly reported "not captured".
    #
    # It is now the mean of the induction and probation experience surveys. Two rules from
    # hrms_survey_service travel with the figure and are the reason it can be shown at all:
    # the aggregation returns SCORES ONLY, never per-respondent rows, and it refuses to
    # return anything below SURVEY_MIN_RESPONSES. A satisfaction survey a manager can
    # de-anonymise measures nothing.
    from app.models.hrms import SURVEY_MIN_RESPONSES, SurveyKind
    from app.services import hrms_survey_service as surveys

    parts, suppressed_kinds, total_responses = [], [], 0
    for kind in (SurveyKind.INDUCTION, SurveyKind.PROBATION):
        agg = await surveys.aggregate(company_id, kind=kind, request_nos=request_nos)
        total_responses += agg.get("responses") or 0
        if agg.get("suppressed"):
            # An instrument with NO responses is not suppressed for privacy, it is simply
            # unanswered -- and telling a reader that "individual answers would be
            # identifiable" when there are no answers is a confusing thing to say about an
            # empty set. Only a non-empty group below the threshold is a suppression.
            if agg.get("responses"):
                suppressed_kinds.append(kind.value)
            continue
        if agg.get("average") is not None:
            parts.append((kind.value, agg["average"], agg["responses"]))

    rates = await surveys.issue_rate(company_id, request_nos=request_nos)
    # Weighted by response count, so twenty induction answers and five probation answers do
    # not each count for half. An unweighted mean of two means is a different number, and a
    # misleading one whenever the groups differ in size.
    weighted = (round(sum(avg * n for _, avg, n in parts) / sum(n for _, _, n in parts), 2)
                if parts else None)
    kpis.append({
        "key": "new_hire_satisfaction",
        "label": "New-hire satisfaction (induction + probation surveys)",
        # A 1-5 mean, NOT a percentage. Every other KPI here is a rate; this one is a score,
        # and `scale_max` is what stops a dashboard rendering 4.3 as "4.3%".
        "value": weighted,
        "scale_max": 5,
        "target": None,
        "numerator": None,
        "denominator": total_responses,
        "eligible_n": total_responses,
        "excluded_n": max(0, (rates.get("issued") or 0) - (rates.get("returned") or 0)),
        "excluded_reason": (
            f'{(rates.get("issued") or 0) - (rates.get("returned") or 0)} survey(s) issued '
            f"and not yet answered." if rates.get("issued") else None),
        "meets_target": None,
        "response_rate": rates.get("response_rate"),
        "by_instrument": [{"kind": k, "average": a, "responses": n} for k, a, n in parts],
        "hint": (f'{rates.get("returned")} of {rates.get("issued")} surveys answered.'
                 if rates.get("issued") else None),
        "reason": (
            None if parts else
            (f"Fewer than {SURVEY_MIN_RESPONSES} responses "
             f'({", ".join(suppressed_kinds)}). Individual answers would be identifiable, '
             f"so no figure is shown." if suppressed_kinds else
             "No survey responses yet. The induction survey is issued when the induction "
             "checklist completes, and the probation one at confirmation.")),
    })

    return {
        "applicable": True,
        "range": {"from": start.strftime("%Y-%m-%d"), "to": end.strftime("%Y-%m-%d")},
        "requisitions": len(reqs),
        "candidates": len(candidates),
        "kpis": kpis,
    }


def _separation_date(profile: dict):
    """When somebody left, from whichever field the company records it in.

    Phase INT-2 added `separation_date` and `status` to the employee profile so the 90-day
    retention KPI has something honest to read. `resigned_on` was already there and is still
    read first-equal, because a company that has been filling it in must not see its
    retention figure change because a new field appeared beside it.

    Deliberately NO separation WORKFLOW was built in this phase. A manually set date is
    enough to make the KPI honest, and a resignation/notice/exit-interview flow is a feature
    with its own approvals and its own consequences -- not a side effect of a reporting fix.
    """
    profile = profile or {}
    date = profile.get("separation_date") or profile.get("resigned_on")
    if date:
        return date
    # A status that means "gone" with no date recorded is a gap, not a departure we can
    # measure. Treated as still employed rather than as a 90-day failure, because guessing
    # a date would put a number nobody chose into a compliance figure.
    return None


async def _client_names(client_ids) -> dict:
    """Current names for a set of client ids, read from Companies.

    Requisitions denormalise `client_name` at write time, which is right for speed but goes
    stale the moment a company is renamed. Under the old HRMS client master a rename fanned
    out across every affected row in a bulk write; clients are companies now, and HRMS does
    not own that write -- so the name is refreshed on READ instead. One query per report, and
    a rename is correct everywhere immediately rather than after a sync step somebody has to
    remember.
    """
    ids = {str(c) for c in client_ids if c}
    if not ids:
        return {}
    from app.services.hrms_client_service import COLL_COMPANIES

    oids = []
    for i in ids:
        try:
            oids.append(ObjectId(i))
        except (InvalidId, TypeError):
            continue
    if not oids:
        return {}
    rows = await get_collection(COLL_COMPANIES).find(
        {"_id": {"$in": oids}}, {"name": 1}).to_list(len(oids))
    return {str(r["_id"]): r.get("name") for r in rows}


async def _client_comparison(actor: dict, company_id: str, start: datetime,
                             end: datetime) -> list:
    """One row per client, for the "all clients" view.

    Bounded work: one requisition read, one candidate read, then arithmetic. The naive
    shape — re-running `dashboard` per client — would issue six reads per client and get
    slower with every client won.
    """
    scope = await _scope(actor, company_id)
    reqs = await get_collection(COLL_REQUISITIONS).find(
        {**scope}, {"request_no": 1, "client_id": 1, "client_name": 1, "vacancy": 1,
                    "closing_status": 1}).to_list(SCAN_CAP)
    if not reqs:
        return []

    by_request = {r["request_no"]: r for r in reqs if r.get("request_no")}
    candidates = await get_collection(COLL_CANDIDATES).find(
        {**scope, **_window("applied_at", start, end)},
        {"uk": 1, "application_status": 1, "request_no": 1, "client_share": 1,
         "client_share_status": 1}).to_list(SCAN_CAP)
    evidence = await _evidence_ranks(scope)

    names = await _client_names(r.get("client_id") for r in reqs)

    buckets = {}
    for r in reqs:
        # Requisitions with no client are grouped under a single explicit bucket rather
        # than dropped: "in-house" is an answer, and silently omitting those rows would make
        # the comparison's total disagree with the dashboard's.
        key = r.get("client_id") or "__none__"
        bucket = buckets.setdefault(key, {
            "client_id": r.get("client_id"),
            "client_name": (names.get(str(r.get("client_id")))
                            or r.get("client_name") or "In-house / no client"),
            "requisitions": 0, "vacancies": 0, "candidates": [], "ranks": [],
        })
        bucket["requisitions"] += 1
        if r.get("closing_status") == ReqClosing.OPEN.value:
            bucket["vacancies"] += int(r.get("vacancy") or 1)

    for c in candidates:
        req = by_request.get(c.get("request_no"))
        key = (req or {}).get("client_id") or "__none__"
        bucket = buckets.get(key)
        if not bucket:
            continue
        bucket["candidates"].append(c)
        bucket["ranks"].append(_effective(c, evidence))

    rows = []
    for bucket in buckets.values():
        metrics = _cv_metrics(bucket["candidates"], bucket["ranks"])
        rows.append({
            "client_id": bucket["client_id"],
            "client_name": bucket["client_name"],
            "requisitions": bucket["requisitions"],
            "vacancies": bucket["vacancies"],
            **metrics,
        })
    rows.sort(key=lambda r: (-r["total"], r["client_name"]))
    return rows


# ─────────────────────────────────────────────────────────────
# Phase 11-R, Item 4 — position-wise CV status matrix
# ─────────────────────────────────────────────────────────────
async def positions(actor: dict, company_id: str, *, date_from: str = None,
                    date_to: str = None, client_id: str = None) -> dict:
    """Rows = requisition, columns = a count for every application status.

    Same `_scope`, same SCAN_CAP, same window validation as everything else here. Read-only.

    The status columns come from AppStatus itself rather than a hand-kept list, so a stage
    added in a later phase appears in this matrix automatically instead of silently
    vanishing from a report somebody trusts.
    """
    start, end = parse_range(date_from, date_to)
    scope = await _scope(actor, company_id, client_id)

    reqs = await get_collection(COLL_REQUISITIONS).find(
        {**scope, **_window("created_at", start, end)},
        {"request_no": 1, "designation_name": 1, "department_name": 1, "vacancy": 1,
         "urgency_level": 1, "closing_status": 1, "approval_status": 1,
         "client_id": 1, "client_name": 1, "requisition_type": 1}).to_list(SCAN_CAP)
    if not reqs:
        return {"range": {"from": start.strftime("%Y-%m-%d"), "to": end.strftime("%Y-%m-%d")},
                "statuses": [s.value for s in AppStatus], "rows": [], "total": 0,
                "scoped_to_own_requisitions": hrms_role(actor) == HrmsRole.MANAGER}

    request_nos = [r["request_no"] for r in reqs if r.get("request_no")]
    candidates = await get_collection(COLL_CANDIDATES).find(
        {**scope, "request_no": {"$in": request_nos}},
        {"uk": 1, "application_status": 1, "request_no": 1, "client_share": 1,
         "client_share_status": 1}).to_list(SCAN_CAP)
    evidence = await _evidence_ranks(scope)

    grouped = {}
    for c in candidates:
        grouped.setdefault(c.get("request_no"), []).append(c)

    names = await _client_names(r.get("client_id") for r in reqs)
    statuses = [s.value for s in AppStatus]
    rows = []
    for r in reqs:
        mine = grouped.get(r.get("request_no"), [])
        ranks = [_effective(c, evidence) for c in mine]
        counts = {s: 0 for s in statuses}
        for c in mine:
            status = c.get("application_status")
            if status in counts:
                counts[status] += 1
        rows.append({
            "request_no": r.get("request_no"),
            "designation": r.get("designation_name"),
            "department": r.get("department_name"),
            "client_name": (names.get(str(r.get("client_id")))
                            or r.get("client_name")),
            "requisition_type": r.get("requisition_type"),
            "vacancy": int(r.get("vacancy") or 1),
            "urgency": r.get("urgency_level"),
            "closing_status": r.get("closing_status"),
            "approval_status": r.get("approval_status"),
            "counts": counts,
            "totals": {"candidates": len(mine), **_cv_metrics(mine, ranks)},
        })

    rows.sort(key=lambda x: -x["totals"]["candidates"])
    return {
        "range": {"from": start.strftime("%Y-%m-%d"), "to": end.strftime("%Y-%m-%d")},
        "statuses": statuses,
        "rows": rows,
        "total": len(rows),
        "scoped_to_own_requisitions": hrms_role(actor) == HrmsRole.MANAGER,
    }


async def _time_to_hire(scope: dict, start: datetime, end: datetime) -> dict:
    """Median and mean days from application to offer acceptance.

    Measured between two first-class timestamps -- `candidates.applied_at` and
    `offers.responded_at` -- rather than by parsing stage-change strings out of the audit
    log. The audit trail records the transitions, but its `detail` is prose written for a
    human; deriving a metric from it would make the number hostage to a wording change.

    Median AND mean, because one slow senior hire skews a mean badly and a reader deserves
    to see both rather than one number chosen for them.
    """
    accepted = await get_collection(COLL_OFFERS).find(
        {**scope, "status": OfferStatus.ACCEPTED.value,
         **_window("responded_at", start, end)},
        {"uk": 1, "responded_at": 1}).to_list(SCAN_CAP)
    if not accepted:
        return {"median_days": None, "mean_days": None, "sample": 0}

    by_uk = {row["uk"]: row["responded_at"] for row in accepted if row.get("uk")}
    applicants = await get_collection(COLL_CANDIDATES).find(
        {"uk": {"$in": list(by_uk)}}, {"uk": 1, "applied_at": 1}).to_list(SCAN_CAP)

    spans = []
    for row in applicants:
        applied, responded = row.get("applied_at"), by_uk.get(row.get("uk"))
        if not applied or not responded:
            continue
        days = (responded - applied).total_seconds() / 86400.0
        if days >= 0:                      # clock skew or a back-dated import
            spans.append(days)

    if not spans:
        return {"median_days": None, "mean_days": None, "sample": 0}
    spans.sort()
    mid = len(spans) // 2
    median = spans[mid] if len(spans) % 2 else (spans[mid - 1] + spans[mid]) / 2
    return {"median_days": round(median, 1),
            "mean_days": round(sum(spans) / len(spans), 1),
            "sample": len(spans)}


# ─────────────────────────────────────────────────────────────
# Funnel
# ─────────────────────────────────────────────────────────────
async def funnel(actor: dict, company_id: str, *, date_from: str = None,
                 date_to: str = None, client_id: str = None) -> dict:
    """The hiring funnel, by effective rank.

    Each stage counts candidates who reached AT LEAST that stage, so the series can never
    increase -- which is what makes stage-to-stage conversion meaningful.
    """
    start, end = parse_range(date_from, date_to)
    scope = await _scope(actor, company_id, client_id)

    candidates = await get_collection(COLL_CANDIDATES).find(
        {**scope, **_window("applied_at", start, end)},
        {"uk": 1, "application_status": 1}).to_list(SCAN_CAP)
    evidence = await _evidence_ranks(scope)
    ranks = [_effective(c, evidence) for c in candidates]

    stages, previous, top = [], None, None
    for key, label, min_rank in FUNNEL_STAGES:
        count = sum(1 for r in ranks if r >= min_rank)
        if top is None:
            top = count
        stages.append({
            "key": key, "label": label, "count": count,
            # Conversion from the PREVIOUS stage, and share of the top of the funnel. Both,
            # because "60% of the previous stage" and "3% of all applicants" answer
            # different questions and showing only one invites the wrong conclusion.
            "from_previous": conversion(count, previous) if previous is not None else 100.0,
            "of_total": conversion(count, top),
        })
        previous = count

    lost = {
        "rejected": sum(1 for c in candidates
                        if c.get("application_status") == AppStatus.REJECTED.value),
        "on_hold": sum(1 for c in candidates
                       if c.get("application_status") == AppStatus.ON_HOLD.value),
        "declined": sum(1 for c in candidates
                        if c.get("application_status") == AppStatus.OFFER_DECLINED.value),
        "duplicate": sum(1 for c in candidates
                         if c.get("application_status") == AppStatus.DUPLICATE.value),
    }

    return {
        "range": {"from": start.strftime("%Y-%m-%d"), "to": end.strftime("%Y-%m-%d")},
        "total": len(candidates),
        "stages": stages,
        "lost": lost,
        # Surfaced so a reader can tell "nobody has been interviewed" from "we could not
        # interpret these rows". A silent zero would look identical.
        "unranked": sum(1 for r in ranks if r == 0),
    }


# ─────────────────────────────────────────────────────────────
# Breakdowns
# ─────────────────────────────────────────────────────────────
async def breakdown(actor: dict, company_id: str, by: str, *, date_from: str = None,
                    date_to: str = None, client_id: str = None) -> dict:
    """Group counts along one allow-listed dimension."""
    spec = BREAKDOWN_FIELDS.get(by)
    if not spec:
        raise HTTPException(status_code=422, detail="Unknown breakdown.")
    collection, field, label = spec

    start, end = parse_range(date_from, date_to)
    scope = await _scope(actor, company_id, client_id)
    date_field = "applied_at" if collection == COLL_CANDIDATES else "created_at"

    rows = await get_collection(collection).aggregate([
        {"$match": {**scope, **_window(date_field, start, end),
                    field: {"$nin": [None, ""]}}},
        {"$group": {"_id": f"${field}", "count": {"$sum": 1}}},
        {"$sort": {"count": -1, "_id": 1}},
        {"$limit": MAX_BREAKDOWN_ROWS},
    ]).to_list(MAX_BREAKDOWN_ROWS)

    total = sum(r["count"] for r in rows)
    return {
        "by": by, "label": label,
        "range": {"from": start.strftime("%Y-%m-%d"), "to": end.strftime("%Y-%m-%d")},
        "total": total,
        "rows": [{"name": r["_id"], "count": r["count"],
                  "share": conversion(r["count"], total)} for r in rows],
        "truncated": len(rows) >= MAX_BREAKDOWN_ROWS,
    }


# ─────────────────────────────────────────────────────────────
# Detailed reports
# ─────────────────────────────────────────────────────────────
def _spec(entity: str) -> dict:
    spec = REPORT_ENTITIES.get(entity)
    if not spec:
        raise HTTPException(status_code=404, detail="Unknown report.")
    return spec


def _project(spec: dict, include_salary: bool) -> list:
    """The columns this caller may see. Compensation columns are DROPPED, not blanked, so a
    reader cannot mistake "you may not see this" for "there is no figure"."""
    return [(key, label) for key, label in spec["columns"]
            if include_salary or key not in SALARY_REPORT_COLUMNS]


def _derive(entity: str, row: dict) -> dict:
    """Fill the DERIVED report columns for one row (Phase 11-R).

    Some columns are computed rather than stored — `budget_status` most notably, which is a
    function of two figures precisely so a correction can never leave a stale flag behind
    (models.budget_status). The report projects a fixed column list off the raw document, so
    without this the column would render empty and read as "no budget recorded".

    Read-only: the row is a copy, and nothing is written back.
    """
    if entity == "requisitions":
        row = dict(row)
        row["budget_status"] = budget_status(row)
        return row
    if entity == "candidates":
        row = dict(row)
        # Prefer the sub-document, fall back to the denormalised field, so rows written at
        # any point since this phase read correctly.
        share = row.get("client_share") or {}
        row["client_share_status"] = share.get("status") or row.get("client_share_status")
        return row
    return row


def _cell(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return ""                       # never flatten a structure into a report cell
    return value


async def _query(actor: dict, company_id: str, entity: str, *, search: str = None,
                 date_from: str = None, date_to: str = None,
                 client_id: str = None) -> tuple:
    spec = _spec(entity)
    start, end = parse_range(date_from, date_to)
    scope = await _scope(actor, company_id, client_id)

    query = {**scope, **_window(spec["date_field"], start, end)}
    if search:
        term = (search or "").strip()[:80]
        if term:
            escaped = re.escape(term)
            query["$or"] = [{f: {"$regex": escaped, "$options": "i"}}
                            for f in spec["search"]]
    return spec, query, (start, end)


async def report(actor: dict, company_id: str, entity: str, *, page: int = 1,
                 page_size: int = None, search: str = None, date_from: str = None,
                 date_to: str = None, client_id: str = None) -> dict:
    """One page of a detailed report."""
    spec, query, (start, end) = await _query(
        actor, company_id, entity, search=search, date_from=date_from, date_to=date_to,
        client_id=client_id)

    page = max(1, int(page or 1))
    page_size = max(1, min(int(page_size or 25), MAX_REPORT_PAGE_SIZE))
    columns = _project(spec, can(actor, Cap.EMPLOYEE_SALARY_READ))

    coll = get_collection(spec["collection"])
    total = await coll.count_documents(query)
    rows = await coll.find(query).sort(spec["date_field"], -1).skip(
        (page - 1) * page_size).limit(page_size).to_list(page_size)

    return {
        "entity": entity,
        "range": {"from": start.strftime("%Y-%m-%d"), "to": end.strftime("%Y-%m-%d")},
        "columns": [{"key": k, "label": lbl} for k, lbl in columns],
        "rows": [{k: _cell(_derive(entity, r).get(k)) for k, _ in columns} for r in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size if total else 0,
        "salary_visible": can(actor, Cap.EMPLOYEE_SALARY_READ),
        # Told to the UI so a hiring manager's table is labelled honestly rather than
        # reading as company-wide.
        "scoped_to_own_requisitions": hrms_role(actor) == HrmsRole.MANAGER,
    }


async def export_rows(actor: dict, company_id: str, entity: str, *, search: str = None,
                      date_from: str = None, date_to: str = None,
                      client_id: str = None) -> dict:
    """Every row for an export, up to the cap, plus an honest truncation flag."""
    spec, query, (start, end) = await _query(
        actor, company_id, entity, search=search, date_from=date_from, date_to=date_to,
        client_id=client_id)

    columns = _project(spec, can(actor, Cap.EMPLOYEE_SALARY_READ))
    coll = get_collection(spec["collection"])
    total = await coll.count_documents(query)
    # One more than the cap, so truncation is detected from the data rather than inferred
    # from a count that could have changed between the two queries.
    rows = await coll.find(query).sort(spec["date_field"], -1).to_list(MAX_EXPORT_ROWS + 1)
    truncated = len(rows) > MAX_EXPORT_ROWS
    rows = rows[:MAX_EXPORT_ROWS]

    return {
        "entity": entity,
        "columns": columns,
        "rows": [[_cell(_derive(entity, r).get(k)) for k, _ in columns] for r in rows],
        "total": total,
        "returned": len(rows),
        "truncated": truncated,
        "range": (start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")),
    }


# ─────────────────────────────────────────────────────────────
# File rendering
# ─────────────────────────────────────────────────────────────
def render_csv(payload: dict) -> bytes:
    """UTF-8 with a BOM, so Excel opens accented names correctly instead of as mojibake."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow([label for _, label in payload["columns"]])
    for row in payload["rows"]:
        writer.writerow(row)
    if payload.get("truncated"):
        writer.writerow([])
        writer.writerow([f"NOTE: truncated to the first {MAX_EXPORT_ROWS} rows of "
                         f"{payload['total']}. Narrow the date range to export the rest."])
    return buf.getvalue().encode("utf-8-sig")


def render_xlsx(payload: dict) -> bytes:
    """openpyxl is imported lazily -- an optional dependency must not cost import time on
    every request, and CSV remains available if it is absent."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=503,
            detail="Excel export is unavailable on this server. Please export CSV.")

    wb = Workbook()
    ws = wb.active
    ws.title = payload["entity"][:31].title()

    headers = [label for _, label in payload["columns"]]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in payload["rows"]:
        ws.append(list(row))

    for i, header in enumerate(headers, start=1):
        widest = max([len(str(header))]
                     + [len(str(r[i - 1])) for r in payload["rows"][:200]] or [0])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(
            max(widest + 2, 10), 48)

    if payload.get("truncated"):
        ws.append([])
        ws.append([f"NOTE: truncated to the first {MAX_EXPORT_ROWS} rows of "
                   f"{payload['total']}. Narrow the date range to export the rest."])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_filename(entity: str, fmt: str, date_range: tuple) -> str:
    return f"hrms_{entity}_{date_range[0]}_to_{date_range[1]}.{fmt}"
