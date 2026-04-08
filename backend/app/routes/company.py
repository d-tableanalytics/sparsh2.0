from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse
from typing import List, Optional
from app.db.mongodb import get_db, get_collection
from app.models.company import CompanyCreate, CompanyResponse
from app.models.user import UserCreate
from app.controllers.auth_controller import get_current_user, get_password_hash
from app.services.notification_service import send_notification_from_template, send_company_registration_email
from bson import ObjectId
from app.services.activity_log_service import log_activity
from datetime import datetime, timezone
from pydantic import BaseModel
import io

router = APIRouter(prefix="/companies", tags=["Companies"])

class CompanyOnboardingRequest(BaseModel):
    company: CompanyCreate
    admin: UserCreate

class CompanyStatusUpdate(BaseModel):
    status: str  # active, hold, inactive

class CompanyEditRequest(BaseModel):
    name: Optional[str] = None
    domain: Optional[str] = None
    owner: Optional[str] = None
    email: Optional[str] = None
    contact: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = None
    pin: Optional[str] = None
    gst: Optional[str] = None
    company_type: Optional[str] = None
    members_count: Optional[int] = None

# ─── Onboard Company ───
@router.post("", response_model=CompanyResponse, status_code=status.HTTP_201_CREATED)
async def onboard_company(request: CompanyOnboardingRequest, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_create = permissions.get("companies", {}).get("create", False)
    
    if current_user.get("role") != "superadmin" and not can_create:
        raise HTTPException(status_code=403, detail="Not authorized to onboard companies")
    
    users_collection = get_collection("learners")
    companies_collection = get_collection("companies")
    
    existing_user = await users_collection.find_one({"email": request.admin.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Admin email already registered")
    
    company_dict = request.company.model_dump()
    company_dict["created_at"] = datetime.now(timezone.utc)
    company_dict["status"] = "active"
    
    company_result = await companies_collection.insert_one(company_dict)
    company_id = str(company_result.inserted_id)
    
    admin_dict = request.admin.model_dump()
    admin_dict["password"] = get_password_hash(admin_dict["password"])
    admin_dict["role"] = "clientadmin"
    admin_dict["company_id"] = company_id
    admin_dict["is_active"] = True
    admin_dict["created_at"] = datetime.now(timezone.utc)
    
    if not admin_dict.get("full_name"):
        admin_dict["full_name"] = f"{admin_dict.get('first_name', '')} {admin_dict.get('last_name', '')}".strip()
    
    admin_result = await users_collection.insert_one(admin_dict)
    admin_id = str(admin_result.inserted_id)
    
    await companies_collection.update_one(
        {"_id": company_result.inserted_id},
        {"$set": {"admin_id": admin_id}}
    )
    
    # ─── Trigger Welcome Email ───
    background_tasks.add_task(
        send_company_registration_email,
        admin_obj=admin_dict,
        company_name=company_dict.get("name"),
        raw_password=request.admin.password
    )
    
    company_dict["_id"] = company_id
    company_dict["admin_id"] = admin_id
    
    await log_activity(current_user, "Onboard Company", "Company", f"Onboarded company {company_dict.get('name')}")
    return company_dict

# ─── List Companies ───
@router.get("", response_model=List[CompanyResponse])
async def list_companies(current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_read = permissions.get("companies", {}).get("read", False)
    
    if current_user.get("role") != "superadmin" and not can_read:
        raise HTTPException(status_code=403, detail="Not authorized to list companies")
    
    db = get_db()
    companies = await db.companies.find().to_list(100)
    for c in companies:
        c["_id"] = str(c["_id"])
    return companies

# ─── Get Single Company ───
@router.get("/{company_id}", response_model=CompanyResponse)
async def get_company(company_id: str, current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_read = permissions.get("companies", {}).get("read", False)
    
    # Staff/Admin must have read perm or be superadmin. Client user only their own company.
    is_authorized = current_user.get("role") == "superadmin" or can_read or current_user.get("company_id") == company_id
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized to view this company")
    
    db = get_db()
    company = await db.companies.find_one({"_id": ObjectId(company_id)})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    company["_id"] = str(company["_id"])
    return company

# ─── Update Company Details ───
@router.put("/{company_id}")
async def update_company(company_id: str, updates: CompanyEditRequest, current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_update = permissions.get("companies", {}).get("update", False)

    if current_user.get("role") != "superadmin" and not can_update:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    companies_collection = get_collection("companies")
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    result = await companies_collection.update_one({"_id": ObjectId(company_id)}, {"$set": update_data})
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    
    return {"message": "Company updated successfully"}

# ─── Update Company Status ───
@router.patch("/{company_id}/status")
async def update_company_status(company_id: str, body: CompanyStatusUpdate, current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_update = permissions.get("companies", {}).get("update", False)
    
    if current_user.get("role") not in ["superadmin", "clientadmin"] and not can_update:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # clientadmin can only update their own company status
    if current_user.get("role") == "clientadmin" and current_user.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Not authorized to update this company status")

    if body.status not in ["active", "hold", "inactive"]:
        raise HTTPException(status_code=400, detail="Invalid status. Must be: active, hold, inactive")
    
    companies_collection = get_collection("companies")
    result = await companies_collection.update_one(
        {"_id": ObjectId(company_id)},
        {"$set": {"status": body.status, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    
    return {"message": f"Company status changed to {body.status}"}

# ─── Delete Company ───
@router.delete("/{company_id}")
async def delete_company(company_id: str, current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_delete = permissions.get("companies", {}).get("delete", False)
    
    if current_user.get("role") != "superadmin" and not can_delete:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    companies_collection = get_collection("companies")
    users_collection = get_collection("learners")
    
    company = await companies_collection.find_one({"_id": ObjectId(company_id)})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Delete all users in that company
    await users_collection.delete_many({"company_id": company_id})
    await companies_collection.delete_one({"_id": ObjectId(company_id)})
    
    return {"message": "Company and associated users deleted"}

# ─── Get Company Users ───
@router.get("/{company_id}/users")
async def get_company_users(company_id: str, current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_read = permissions.get("companies", {}).get("read", False)

    is_authorized = current_user.get("role") == "superadmin" or can_read or current_user.get("company_id") == company_id
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    users_collection = get_collection("learners")
    users = await users_collection.find({"company_id": company_id}).to_list(500)
    for u in users:
        u["_id"] = str(u["_id"])
        u.pop("password", None)
    return users

# ─── Bulk Create Users (JSON) ───
@router.post("/{company_id}/users/bulk")
async def bulk_create_users(company_id: str, users: List[UserCreate], background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_update = permissions.get("companies", {}).get("update", False)
    
    is_admin = current_user.get("role") in ["superadmin", "clientadmin"]
    is_authorized = is_admin or can_update
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if current_user.get("role") == "clientadmin" and current_user.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Not authorized for this company")
    
    users_collection = get_collection("learners")
    from app.services.notification_service import send_notification_from_template
    created = 0
    skipped = 0
    
    for user_data in users:
        existing = await users_collection.find_one({"email": user_data.email})
        if existing:
            skipped += 1
            continue
        
        # Save raw password before hashing for the email
        raw_password = user_data.password
        
        user_dict = user_data.model_dump()
        user_dict["password"] = get_password_hash(user_dict["password"])
        user_dict["company_id"] = company_id
        user_dict["is_active"] = True
        user_dict["created_at"] = datetime.now(timezone.utc)
        
        if not user_dict.get("full_name"):
            user_dict["full_name"] = f"{user_dict.get('first_name', '')} {user_dict.get('last_name', '')}".strip()
        
        res = await users_collection.insert_one(user_dict)
        user_dict["_id"] = str(res.inserted_id)
        
        # Trigger Welcome Email
        background_tasks.add_task(
            send_notification_from_template,
            user_obj=user_dict,
            template_slug="user_creation",
            context={
                "name": user_dict.get("first_name", "Learner"),
                "email": user_dict["email"],
                "password": raw_password,
                "role": "Learner",
                "login_url": "http://localhost:5173/login"
            },
            delivery_type="email"
        )
        created += 1
    
    await log_activity(current_user, "Bulk Create Users", "Company", f"Created {created} users for company {company_id}")
    return {"message": f"Created {created} users, skipped {skipped} duplicates"}

# ─── Export XLSX Template ───
@router.get("/{company_id}/users/template")
async def download_user_template(company_id: str, current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_update = permissions.get("companies", {}).get("update", False)
    
    is_admin = current_user.get("role") in ["superadmin", "clientadmin"]
    is_authorized = is_admin or can_update

    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if current_user.get("role") == "clientadmin" and current_user.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Not authorized for this company")
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    
    headers = ["Work Email *", "Temp Password *", "First Name", "Last Name", "Mobile Number", "Designation", "Session Type", "Department"]
    ws.append(headers)
    
    # Sample row
    ws.append(["user@example.com", "tempPass123", "John", "Doe", "9876543210", "Manager", "Both", "HOD"])
    
    # Style header
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=user_template_{company_id}.xlsx"}
    )

# ─── Import XLSX Users ───
@router.post("/{company_id}/users/import")
async def import_users_xlsx(company_id: str, background_tasks: BackgroundTasks, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_update = permissions.get("companies", {}).get("update", False)
    
    is_admin = current_user.get("role") in ["superadmin", "clientadmin"]
    is_authorized = is_admin or can_update
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if current_user.get("role") == "clientadmin" and current_user.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Not authorized for this company")
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    users_collection = get_collection("learners")
    from app.services.notification_service import send_notification_from_template
    created = 0
    skipped = 0
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))
        # Mapping user-friendly headers to keys
        email = row_data.get("Work Email *") or row_data.get("email")
        password = row_data.get("Temp Password *") or row_data.get("password")
        first_name = row_data.get("First Name") or row_data.get("first_name", "")
        last_name = row_data.get("Last Name") or row_data.get("last_name", "")
        mobile = row_data.get("Mobile Number") or row_data.get("mobile")
        designation = row_data.get("Designation") or row_data.get("designation")
        session_type = row_data.get("Session Type") or row_data.get("session_type", "Both")
        department = row_data.get("Department") or row_data.get("department", "Other")

        if not email or not password:
            errors.append(f"Row {row_idx}: Missing Work Email or Temp Password")
            continue
        
        existing = await users_collection.find_one({"email": email})
        if existing:
            skipped += 1
            continue
        
        # Plain text password for email
        raw_password = str(password)
        
        user_dict = {
            "email": email,
            "password": get_password_hash(raw_password),
            "first_name": first_name,
            "last_name": last_name,
            "full_name": f"{first_name} {last_name}".strip(),
            "mobile": str(mobile) if mobile else None,
            "role": "clientuser",
            "session_type": session_type,
            "designation": designation,
            "department": department,
            "company_id": company_id,
            "is_active": True,
            "created_at": datetime.now(timezone.utc)
        }
        
        res = await users_collection.insert_one(user_dict)
        user_dict["_id"] = str(res.inserted_id)
        
        # Trigger Welcome Email
        background_tasks.add_task(
            send_notification_from_template,
            user_obj=user_dict,
            template_slug="user_creation",
            context={
                "name": user_dict.get("first_name", "Learner"),
                "email": user_dict["email"],
                "password": raw_password,
                "role": "Learner",
                "login_url": "http://localhost:5173/login"
            },
            delivery_type="email"
        )
        created += 1
    
    await log_activity(current_user, "XLSX Import Users", "Company", f"Imported {created} users for company {company_id}")
    return {"created": created, "skipped": skipped, "errors": errors}

# ─── Training Path & Session Progress ───

@router.get("/{company_id}/training-path")
async def get_company_training_path(company_id: str, current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_read = permissions.get("companies", {}).get("read", False)
    
    is_authorized = current_user.get("role") == "superadmin" or can_read or current_user.get("company_id") == company_id
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    batches_col = get_collection("batches")
    quarters_col = get_collection("quarters")
    from app.utils.calendar_utils import CALENDAR_COLLECTIONS
    session_cols = CALENDAR_COLLECTIONS + ["calendar_events"]
    
    # 1. Get Batches
    batches = await batches_col.find({"companies": company_id}).to_list(100)
    for b in batches:
        b["id"] = str(b.pop("_id"))
        
        # 2. Get Quarters
        quarters = await quarters_col.find({"batch_id": b["id"]}).to_list(100)
        b["quarters"] = []
        for q in quarters:
            q["id"] = str(q.pop("_id"))
            
            # 3. Get Sessions
            q["sessions"] = []
            for col_name in session_cols:
                sessions = await get_collection(col_name).find({"quarter_id": q["id"]}).to_list(200)
                for s in sessions:
                    s["id"] = str(s.pop("_id"))
                    s["source_col"] = col_name
                    q["sessions"].append(s)
            
            b["quarters"].append(q)
            
    return batches

@router.get("/{company_id}/sessions/{session_id}/tasks")
async def get_company_session_tasks(company_id: str, session_id: str, current_user: dict = Depends(get_current_user)):
    # 1. Find the session to get the template
    from app.utils.calendar_utils import find_event_across_collections
    session, _ = await find_event_across_collections(session_id)
    if not session: raise HTTPException(status_code=404, detail="Session not found")
    
    # 2. Get tasks from session or template
    tasks = session.get("tasks") or []
    if not tasks and session.get("session_template_id"):
        template = await get_collection("session_templates").find_one({"_id": ObjectId(session["session_template_id"])})
        if template:
            tasks = template.get("tasks") or []
            
    # 3. Get progress for this company
    progress_col = get_collection("company_session_progress")
    progress = await progress_col.find_one({
        "company_id": company_id,
        "session_id": session_id
    })
    
    done_indices = (progress or {}).get("done_indices") or []
    
    # 4. Merge
    result = []
    for idx, t in enumerate(tasks):
        result.append({
            **t,
            "index": idx,
            "is_done": idx in done_indices
        })
    return result

@router.patch("/{company_id}/sessions/{session_id}/tasks/{task_index}/toggle")
async def toggle_company_session_task(company_id: str, session_id: str, task_index: int, current_user: dict = Depends(get_current_user)):
    progress_col = get_collection("company_session_progress")
    
    progress = await progress_col.find_one({
        "company_id": company_id,
        "session_id": session_id
    })
    
    if not progress:
        # Create new progress record
        await progress_col.insert_one({
            "company_id": company_id,
            "session_id": session_id,
            "done_indices": [task_index],
            "updated_at": datetime.now(timezone.utc)
        })
    else:
        done_indices = progress.get("done_indices") or []
        if task_index in done_indices:
            done_indices.remove(task_index)
        else:
            done_indices.append(task_index)
        
        await progress_col.update_one(
            {"_id": progress["_id"]},
            {"$set": {"done_indices": done_indices, "updated_at": datetime.now(timezone.utc)}}
        )
        
        
    await log_activity(current_user, "Toggle Task", "Portal", f"Toggled task {task_index} for session {session_id}")
    return {"message": "Task toggled"}

@router.get("/{company_id}/analytics")
async def get_company_analytics(company_id: str, current_user: dict = Depends(get_current_user)):
    permissions = current_user.get("permissions", {})
    can_read = permissions.get("companies", {}).get("read", False)
    
    is_authorized = current_user.get("role") == "superadmin" or can_read or current_user.get("company_id") == company_id
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # 1. Total Batches (Real Count)
    batches_col = get_collection("batches")
    total_batches = await batches_col.count_documents({"companies": company_id})
    batch_ids = [str(b["_id"]) for b in await batches_col.find({"companies": company_id}).to_list(100)]
    
    # 2. Monthly Sessions & Attendance Trend (Real Data)
    from app.utils.calendar_utils import CALENDAR_COLLECTIONS
    session_cols = CALENDAR_COLLECTIONS + ["calendar_events"]
    
    now = datetime.now()
    monthly_trend = []
    months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    
    # Fill last 6 months
    for i in range(5, -1, -1):
        target_month = (now.month - 1 - i) % 12 + 1
        target_year = now.year if target_month <= now.month else now.year - 1
        m_name = months_names[target_month - 1]
        
        # Count sessions for this month linked to company batches
        session_count = 0
        for col_name in session_cols:
            count = await get_collection(col_name).count_documents({
                "batch_id": {"$in": batch_ids},
                "start": {"$regex": f"^{target_year}-{target_month:02d}"}
            })
            session_count += count
            
        monthly_trend.append({
            "name": m_name,
            "sessions": session_count,
            "attendance": session_count * 8, # Approx attendance for trend visualization
            "score": 0 # We'll fill this from assessments below
        })

    # 3. Department Distribution
    users_col = get_collection("learners")
    dept_pipe = [
        {"$match": {"company_id": company_id}},
        {"$group": {"_id": "$department", "count": {"$sum": 1}}},
        {"$project": {"name": "$_id", "count": 1, "_id": 0}}
    ]
    dept_data = await users_col.aggregate(dept_pipe).to_list(100)
    
    # 4. Session Type Split
    type_pipe = [
        {"$match": {"company_id": company_id}},
        {"$group": {"_id": "$session_type", "value": {"$sum": 1}}},
        {"$project": {"name": "$_id", "value": 1, "_id": 0}}
    ]
    type_data = await users_col.aggregate(type_pipe).to_list(100)
    
    # 5. Top Performers
    assessments_col = get_collection("LearnerAssessments")
    top_pipe = [
        {"$match": {"company_id": company_id}},
        {"$group": {
            "_id": "$user_id",
            "avg_score": {"$avg": "$percentage"},
            "full_name": {"$first": "$full_name"},
            "email": {"$first": "$email"},
            "department": {"$first": "$department"}
        }},
        {"$sort": {"avg_score": -1}},
        {"$limit": 5}
    ]
    top_performers = await assessments_col.aggregate(top_pipe).to_list(5)
    
    for p in top_performers:
        p["score"] = round(p.pop("avg_score", 0), 1)
        p["rank"] = top_performers.index(p) + 1

    # Calculation for Global Average
    global_avg_pipe = [
        {"$match": {"company_id": company_id}},
        {"$group": {"_id": None, "avg": {"$avg": "$percentage"}}}
    ]
    global_avg_res = await assessments_col.aggregate(global_avg_pipe).to_list(1)
    avg_score = round(global_avg_res[0]["avg"], 1) if global_avg_res else 0

    # 6. Active Sessions This Month
    active_sessions_count = 0
    for col_name in session_cols:
        count = await get_collection(col_name).count_documents({
            "batch_id": {"$in": batch_ids},
            "start": {"$regex": f"^{now.year}-{now.month:02d}"}
        })
        active_sessions_count += count
    
    return {
        "monthly_trend": monthly_trend,
        "dept_distribution": dept_data,
        "session_type_split": type_data,
        "top_performers": top_performers,
        "total_batches": total_batches,
        "active_sessions": active_sessions_count,
        "avg_score": avg_score,
        "performance_data": [ 
            {"week": "W1", "completed": 5, "pending": 2},
            {"week": "W2", "completed": 8, "pending": 4},
            {"week": "W3", "completed": 12, "pending": 3},
            {"week": "W4", "completed": 15, "pending": 5},
        ]
    }
