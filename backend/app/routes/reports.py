"""
Admin Reports & Analytics — API router (Super Admin + Admin).

All endpoints are read-only aggregations over existing task/user/activity data.
Nothing here mutates state or changes existing behaviour. Access is gated to the
`superadmin` and `admin` roles via check_role
(see docs/ENTERPRISE_REPORTS_ANALYSIS.md).
"""
import csv
import io
from typing import Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from app.db.mongodb import get_collection
from app.controllers.auth_controller import check_role
from app.utils.calendar_utils import find_event_across_collections
from app.services import report_service as rs
from app.services import report_company_service as rcs
from app.services import report_lms_service as rls

router = APIRouter(prefix="/reports", tags=["Reports"])

# Super Admin + Admin for every route in this module.
admin_only = check_role(["superadmin", "admin"])


def _sort_rows(rows, sort_key: Optional[str], order: str):
    if not sort_key:
        return rows
    reverse = (order or "desc").lower() != "asc"
    if sort_key == "name":
        return sorted(rows, key=lambda r: (r.get("name") or "").lower(), reverse=reverse)
    return sorted(rows, key=lambda r: (r.get(sort_key) if r.get(sort_key) is not None else -1), reverse=reverse)


@router.get("/enterprise-overview")
async def reports_enterprise_overview(
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    department: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    """Executive KPI band: companies, users, coaches, learners, sessions, batches,
    courses (session templates), tasks, assessment score, attendance, completion."""
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    if department:
        tasks = _filter_by_department(tasks, users, department)
    return await rs.compute_enterprise_overview(tasks, users)


@router.get("/overview")
async def reports_overview(
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    department: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    if department:
        tasks = _filter_by_department(tasks, users, department)
    return rs.compute_overview(tasks, users)


@router.get("/company")
async def reports_company(
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    department: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    if department:
        tasks = _filter_by_department(tasks, users, department)
    return rs.compute_company(tasks, users)


@router.get("/departments")
async def reports_departments(
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    return {"departments": rs.compute_company(tasks, users)["departments"]}


@router.get("/doers")
async def reports_doers(
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    department: Optional[str] = None,
    search: Optional[str] = None,
    sort: Optional[str] = Query("score"),
    order: Optional[str] = Query("desc"),
    skip: int = 0,
    limit: int = 25,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    rows = rs.compute_doers(tasks, users)

    # Attach each employee's company name (for the Employee-Wise "Company" column).
    company_names = await rcs._company_name_map()
    for r in rows:
        cid = users.get(r["id"], {}).get("company_id")
        r["company"] = company_names.get(cid) if cid else None

    if department:
        rows = [r for r in rows if r["department"] == department]
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower() or s in (r.get("email") or "").lower()
                or s in (r.get("company") or "").lower()]

    rows = _sort_rows(rows, sort, order)
    total = len(rows)
    # keep 1-based rank stable across pagination
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    page = rows[skip: skip + limit] if limit else rows
    return {"items": page, "total": total, "skip": skip, "limit": limit}


@router.get("/employees-wide")
async def reports_employees_wide(
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    department: Optional[str] = None,
    company_id: Optional[str] = None,
    status: Optional[str] = None,           # active | inactive
    search: Optional[str] = None,
    sort: Optional[str] = Query("score"),
    order: Optional[str] = Query("desc"),
    skip: int = 0,
    limit: int = 25,
    current_user: dict = Depends(admin_only),
):
    """Comprehensive employee/learner report — includes users with zero tasks (task doers
    were the old blind spot). Real task + attendance/session + assessment + activity data."""
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    rows = await rs.compute_employees_wide(tasks, users)

    if department:
        rows = [r for r in rows if r["department"] == department]
    if company_id:
        rows = [r for r in rows if r.get("companyId") == company_id]
    if status:
        want_active = status.lower() == "active"
        rows = [r for r in rows if bool(r["isActive"]) == want_active]
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower()
                or s in (r.get("email") or "").lower()
                or s in (r.get("company") or "").lower()
                or s in (str(r.get("employeeId") or "")).lower()]

    rows = _sort_rows(rows, sort, order)
    total = len(rows)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    page = rows[skip: skip + limit] if limit else rows
    return {"items": page, "total": total, "skip": skip, "limit": limit}


@router.get("/activity")
async def reports_activity(
    department: Optional[str] = None,
    company_id: Optional[str] = None,
    search: Optional[str] = None,
    sort: Optional[str] = Query("lastActivity"),
    order: Optional[str] = Query("desc"),
    skip: int = 0,
    limit: int = 25,
    current_user: dict = Depends(admin_only),
):
    """Activity report — real login/usage from activity_logs + attendance/assessments."""
    users = await rs.load_users()
    result = await rs.compute_activity(users)
    rows = result["items"]
    if department:
        rows = [r for r in rows if r["department"] == department]
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower()
                or s in (r.get("email") or "").lower() or s in (r.get("company") or "").lower()]
    rows = _sort_rows(rows, sort, order)
    total = len(rows)
    page = rows[skip: skip + limit] if limit else rows
    return {"summary": result["summary"], "items": page, "total": total, "skip": skip, "limit": limit}


@router.get("/sessions")
async def reports_sessions(
    search: Optional[str] = None,
    status: Optional[str] = None,
    sort: Optional[str] = Query("date"),
    order: Optional[str] = Query("desc"),
    skip: int = 0,
    limit: int = 25,
    current_user: dict = Depends(admin_only),
):
    """Session report — LMS sessions with attendance % and duration."""
    users = await rs.load_users()
    result = await rs.compute_sessions(users)
    rows = result["items"]
    if status:
        rows = [r for r in rows if (r.get("status") or "").lower() == status.lower()]
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower()]
    rows = _sort_rows(rows, sort, order)
    total = len(rows)
    page = rows[skip: skip + limit] if limit else rows
    return {"summary": result["summary"], "monthly": result["monthly"], "items": page, "total": total, "skip": skip, "limit": limit}


@router.get("/doers/{doer_id}")
async def reports_doer_detail(
    doer_id: str,
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    if doer_id not in users:
        raise HTTPException(status_code=404, detail="Employee not found")
    return rs.compute_doer_detail(tasks, doer_id, users)


@router.get("/doers/{doer_id}/history")
async def reports_doer_history(
    doer_id: str,
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 25,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    rows = rs.doer_task_rows(tasks, doer_id, users)

    if status:
        rows = [r for r in rows if r["status"] == status]
    if priority:
        rows = [r for r in rows if r["priority"] == priority]
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["title"] or "").lower()]

    total = len(rows)
    page = rows[skip: skip + limit] if limit else rows
    return {"items": page, "total": total, "skip": skip, "limit": limit}


@router.get("/doers/{doer_id}/timeline")
async def reports_doer_timeline(
    doer_id: str,
    task_id: str = Query(...),
    current_user: dict = Depends(admin_only),
):
    doc, col = await find_event_across_collections(task_id)
    if not doc or doc.get("type") != "task":
        raise HTTPException(status_code=404, detail="Task not found")
    events = await rs.build_timeline(task_id, col, doc)
    return {"taskId": task_id, "title": doc.get("title"), "events": events}


# --------------------------------------------------------------------------- #
# E2 — Companies                                                               #
# --------------------------------------------------------------------------- #
@router.get("/companies")
async def reports_companies(
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    search: Optional[str] = None,
    sort: Optional[str] = Query("score"),
    order: Optional[str] = Query("desc"),
    skip: int = 0,
    limit: int = 25,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    rows = await rcs.compute_companies(tasks, users)
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower()]
    rows = _sort_rows(rows, sort, order)
    total = len(rows)
    page = rows[skip: skip + limit] if limit else rows
    return {"items": page, "total": total, "skip": skip, "limit": limit}


_COMPANY_EXPORT_COLUMNS = [
    ("rank", "Rank"), ("name", "Company"), ("status", "Status"), ("employees", "Total Employees"),
    ("assigned", "Total Tasks"), ("completed", "Completed"), ("pending", "Pending"), ("overdue", "Overdue"),
    ("attendanceRate", "Attendance %"), ("avgAssessment", "Assessment %"), ("completionRate", "Completion %"),
    ("sessions", "Sessions"), ("score", "Productivity %"), ("rating", "Rating"),
]


@router.get("/companies/export")
async def reports_companies_export(
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    search: Optional[str] = None,
    sort: Optional[str] = Query("score"),
    order: Optional[str] = Query("desc"),
    current_user: dict = Depends(admin_only),
):
    """Export the Company-wise report (CSV / Excel / PDF), respecting filters."""
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    rows = await rcs.compute_companies(tasks, users)
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower()]
    rows = _sort_rows(rows, sort, order)
    for i, r in enumerate(rows):
        r["rank"] = i + 1

    if format == "csv":
        return _export_csv(rows, columns=_COMPANY_EXPORT_COLUMNS, filename="company_report.csv")
    if format == "xlsx":
        return _export_xlsx(rows, columns=_COMPANY_EXPORT_COLUMNS, filename="company_report.xlsx", sheet_title="Companies")
    return _export_pdf(rows, columns=_COMPANY_EXPORT_COLUMNS, filename="company_report.pdf",
                       title="Company-wise Report", subtitle=f"Companies: {len(rows)}")


@router.get("/companies/{company_id}")
async def reports_company_dashboard(
    company_id: str,
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    return await rcs.compute_company_dashboard(company_id, tasks, users)


@router.get("/companies/{company_id}/employees")
async def reports_company_employees(
    company_id: str,
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    search: Optional[str] = None,
    sort: Optional[str] = Query("score"),
    order: Optional[str] = Query("desc"),
    skip: int = 0,
    limit: int = 25,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    att = await rcs._attendance_by_user([uid for uid, u in users.items() if u.get("company_id") == company_id])
    assess = await rcs._assessments_by_user([uid for uid, u in users.items() if u.get("company_id") == company_id])
    rows = rcs.compute_company_employees(company_id, tasks, users, att, assess)
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower() or s in (r.get("email") or "").lower()]
    rows = _sort_rows(rows, sort, order)
    total = len(rows)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    page = rows[skip: skip + limit] if limit else rows
    return {"items": page, "total": total, "skip": skip, "limit": limit}


_COMPANY_EMP_EXPORT_COLUMNS = [
    ("rank", "Rank"), ("name", "Employee"), ("email", "Email"), ("department", "Department"),
    ("assigned", "Assigned"), ("completed", "Completed"), ("pending", "Pending"), ("overdue", "Overdue"),
    ("attendanceRate", "Attendance %"), ("avgAssessment", "Assessment %"), ("completionRate", "Completion %"),
    ("score", "Productivity %"), ("rating", "Rating"),
]


@router.get("/companies/{company_id}/employees/export")
async def reports_company_employees_export(
    company_id: str,
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    search: Optional[str] = None,
    sort: Optional[str] = Query("score"),
    order: Optional[str] = Query("desc"),
    current_user: dict = Depends(admin_only),
):
    """Export one company's employees (CSV / Excel / PDF), respecting filters."""
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    scoped = [uid for uid, u in users.items() if u.get("company_id") == company_id]
    att = await rcs._attendance_by_user(scoped)
    assess = await rcs._assessments_by_user(scoped)
    rows = rcs.compute_company_employees(company_id, tasks, users, att, assess)
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower() or s in (r.get("email") or "").lower()]
    rows = _sort_rows(rows, sort, order)
    for i, r in enumerate(rows):
        r["rank"] = i + 1

    company = await get_collection("companies").find_one({"_id": ObjectId(company_id)}) if ObjectId.is_valid(company_id) else None
    cname = (company or {}).get("name", "company")
    base = f"employee_report_{_safe_filename(cname)}"
    if format == "csv":
        return _export_csv(rows, columns=_COMPANY_EMP_EXPORT_COLUMNS, filename=f"{base}.csv")
    if format == "xlsx":
        return _export_xlsx(rows, columns=_COMPANY_EMP_EXPORT_COLUMNS, filename=f"{base}.xlsx", sheet_title="Employees")
    return _export_pdf(rows, columns=_COMPANY_EMP_EXPORT_COLUMNS, filename=f"{base}.pdf",
                       title=f"Employee-wise Report — {cname}", subtitle=f"Employees: {len(rows)}")


# --------------------------------------------------------------------------- #
# LMS (= Batch) reporting                                                      #
# --------------------------------------------------------------------------- #
@router.get("/lms")
async def reports_lms_list(
    company_id: Optional[str] = None,
    search: Optional[str] = None,
    sort: Optional[str] = Query("completionRate"),
    order: Optional[str] = Query("desc"),
    skip: int = 0,
    limit: int = 25,
    current_user: dict = Depends(admin_only),
):
    users = await rs.load_users()
    rows = await rls.compute_lms_list(users, company_id=company_id)
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower()]
    rows = _sort_rows(rows, sort, order)
    total = len(rows)
    page = rows[skip: skip + limit] if limit else rows
    return {"items": page, "total": total, "skip": skip, "limit": limit}


@router.get("/lms/{batch_id}")
async def reports_lms_dashboard(
    batch_id: str,
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    return await rls.compute_lms_dashboard(batch_id, tasks, users)


@router.get("/lms/{batch_id}/employees")
async def reports_lms_employees(
    batch_id: str,
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    search: Optional[str] = None,
    sort: Optional[str] = Query("score"),
    order: Optional[str] = Query("desc"),
    skip: int = 0,
    limit: int = 25,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    rows = await rls.compute_lms_employees(batch_id, tasks, users)
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower() or s in (r.get("email") or "").lower()]
    rows = _sort_rows(rows, sort, order)
    total = len(rows)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    page = rows[skip: skip + limit] if limit else rows
    return {"items": page, "total": total, "skip": skip, "limit": limit}


# --------------------------------------------------------------------------- #
# E3 — Employee                                                                #
# --------------------------------------------------------------------------- #
@router.get("/employees/{user_id}")
async def reports_employee(
    user_id: str,
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    report = await rcs.compute_employee_report(user_id, tasks, users)
    if report is None:
        raise HTTPException(status_code=404, detail="Employee not found")
    return report


@router.get("/employees/{user_id}/assignments")
async def reports_employee_assignments(
    user_id: str,
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 25,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    rows = rs.doer_task_rows(tasks, user_id, users)
    if status:
        rows = [r for r in rows if r["status"] == status]
    if priority:
        rows = [r for r in rows if r["priority"] == priority]
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["title"] or "").lower()]
    total = len(rows)
    page = rows[skip: skip + limit] if limit else rows
    return {"items": page, "total": total, "skip": skip, "limit": limit}


@router.get("/employees/{user_id}/assessments")
async def reports_employee_assessments(user_id: str, current_user: dict = Depends(admin_only)):
    return {"items": await rcs.employee_assessments(user_id)}


@router.get("/employees/{user_id}/attendance")
async def reports_employee_attendance(user_id: str, current_user: dict = Depends(admin_only)):
    return {"items": await rcs.employee_attendance(user_id)}


@router.get("/employees/{user_id}/timeline")
async def reports_employee_timeline(
    user_id: str,
    task_id: str = Query(...),
    current_user: dict = Depends(admin_only),
):
    doc, col = await find_event_across_collections(task_id)
    if not doc or doc.get("type") != "task":
        raise HTTPException(status_code=404, detail="Task not found")
    events = await rs.build_timeline(task_id, col, doc)
    return {"taskId": task_id, "title": doc.get("title"), "events": events}


# --------------------------------------------------------------------------- #
# Export (CSV / Excel / PDF)                                                   #
# --------------------------------------------------------------------------- #
_DOER_EXPORT_COLUMNS = [
    ("rank", "Rank"), ("name", "Employee"), ("department", "Department"),
    ("assigned", "Assigned"), ("completed", "Completed"), ("pending", "Pending"),
    ("overdue", "Overdue"), ("completionRate", "Completion %"),
    ("score", "Score"), ("avgCompletionDays", "Avg Days"), ("rating", "Rating"),
]


@router.get("/export")
async def reports_export(
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    department: Optional[str] = None,
    lms: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    # Respect the LMS (batch) filter when present, else the org-wide doer set.
    if lms:
        rows = await rls.compute_lms_employees(lms, tasks, users)
    else:
        rows = rs.compute_doers(tasks, users)
    if department:
        rows = [r for r in rows if r["department"] == department]
    for i, r in enumerate(rows):
        r["rank"] = i + 1

    if format == "csv":
        return _export_csv(rows)
    if format == "xlsx":
        return _export_xlsx(rows, sheet_title="Employee Performance")
    return _export_pdf(rows, subtitle=f"Period: {period or 'all_time'} &nbsp;&nbsp; Employees: {len(rows)}")


_EMP_ASSIGNMENT_COLUMNS = [
    ("title", "Assignment"), ("module", "Module"), ("assignedBy", "Assigned By"),
    ("assignedDate", "Assigned"), ("dueDate", "Due"), ("completedDate", "Completed"),
    ("statusLabel", "Status"), ("priority", "Priority"), ("score", "Score"),
]


_EMP_WIDE_COLUMNS = [
    ("rank", "Rank"), ("name", "Employee"), ("email", "Email"), ("employeeId", "Emp ID"),
    ("company", "Company"), ("department", "Department"), ("designation", "Designation"),
    ("assigned", "Task Assigned"), ("completed", "Task Completed"), ("pending", "Pending"), ("overdue", "Overdue"),
    ("totalSessions", "Total Sessions"), ("sessionsAttended", "Attended"), ("sessionsMissed", "Missed"),
    ("attendanceRate", "Attendance %"), ("avgAssessment", "Assessment %"), ("completionRate", "Completion %"),
    ("totalLogins", "Logins"), ("lastLogin", "Last Login"), ("lastActivity", "Last Activity"), ("rating", "Status"),
]


@router.get("/employees-wide/export")
async def reports_employees_wide_export(
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    department: Optional[str] = None,
    company_id: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    """Export the comprehensive employee/learner report (CSV / Excel / PDF), respecting filters."""
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    rows = await rs.compute_employees_wide(tasks, users)

    if department:
        rows = [r for r in rows if r["department"] == department]
    if company_id:
        rows = [r for r in rows if r.get("companyId") == company_id]
    if status:
        want_active = status.lower() == "active"
        rows = [r for r in rows if bool(r["isActive"]) == want_active]
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r["name"] or "").lower()
                or s in (r.get("email") or "").lower() or s in (r.get("company") or "").lower()]
    for i, r in enumerate(rows):
        r["rank"] = i + 1
        r["lastLogin"] = str(r.get("lastLogin") or "")
        r["lastActivity"] = str(r.get("lastActivity") or "")

    if format == "csv":
        return _export_csv(rows, columns=_EMP_WIDE_COLUMNS, filename="employee_report.csv")
    if format == "xlsx":
        return _export_xlsx(rows, columns=_EMP_WIDE_COLUMNS, filename="employee_report.xlsx", sheet_title="Employees")
    return _export_pdf(rows, columns=_EMP_WIDE_COLUMNS, filename="employee_report.pdf",
                       title="Employee Report", subtitle=f"Employees: {len(rows)}")


def _safe_filename(name: str) -> str:
    keep = "".join(c if (c.isalnum() or c in "-_") else "_" for c in (name or "employee"))
    return keep.strip("_") or "employee"


@router.get("/employees/{user_id}/export")
async def reports_employee_export(
    user_id: str,
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    period: Optional[str] = None,
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    """Export a single employee's assignment history (CSV / Excel / PDF)."""
    start_iso, end_iso = rs.period_range(period, startDate, endDate)
    users = await rs.load_users()
    if user_id not in users:
        raise HTTPException(status_code=404, detail="Employee not found")
    tasks = await rs.fetch_tasks(start_iso, end_iso)
    rows = rs.doer_task_rows(tasks, user_id, users)

    name = users[user_id].get("name", "employee")
    base = f"employee_report_{_safe_filename(name)}"
    if format == "csv":
        return _export_csv(rows, columns=_EMP_ASSIGNMENT_COLUMNS, filename=f"{base}.csv")
    if format == "xlsx":
        return _export_xlsx(rows, columns=_EMP_ASSIGNMENT_COLUMNS, filename=f"{base}.xlsx", sheet_title="Assignments")
    return _export_pdf(
        rows, columns=_EMP_ASSIGNMENT_COLUMNS, filename=f"{base}.pdf",
        title=f"Employee Report — {name}",
        subtitle=f"Assignments: {len(rows)}",
    )


def _cell(r, key):
    v = r.get(key)
    return "" if v is None else v


def _export_csv(rows, columns=_DOER_EXPORT_COLUMNS, filename="employee_performance.csv"):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _, label in columns])
    for r in rows:
        writer.writerow([_cell(r, key) for key, _ in columns])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _export_xlsx(rows, columns=_DOER_EXPORT_COLUMNS, filename="employee_performance.xlsx", sheet_title="Report"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet-name limit
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, (_, label) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.fill = header_fill
        c.font = header_font
        ws.column_dimensions[c.column_letter].width = 18
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell(r, key))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _export_pdf(rows, columns=_DOER_EXPORT_COLUMNS, filename="employee_performance.pdf",
                title="Employee Performance Report", subtitle=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), topMargin=18 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 8 * mm))

    data = [[label for _, label in columns]]
    for r in rows:
        data.append([str(_cell(r, key)) for key, _ in columns])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
        ("ALIGN", (3, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# --------------------------------------------------------------------------- #
def _filter_by_department(tasks, users, department):
    def in_dept(doc):
        return any(users.get(did, {}).get("department") == department for did in rs.doer_ids(doc))
    return [t for t in tasks if in_dept(t)]
